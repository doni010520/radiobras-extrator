"""
RADIOBRAS — Extrator Web
Flask app: relatório analítico (xlsx) + download de arquivos (ZIP).
"""

import io
import logging
import os
import re
import sys
import threading
import time
import traceback
import uuid
from datetime import datetime, timedelta

try:
    from zoneinfo import ZoneInfo
    _TZ = ZoneInfo("America/Sao_Paulo")
except Exception:
    _TZ = None

from flask import (Flask, jsonify, render_template, request, send_file,
                   session, redirect, url_for, Response)

sys.path.insert(0, os.path.dirname(__file__))

from extrator_pacientes_analitico import (
    discover_tokens_and_cookies,
    get_credentials,
    parse_html_to_df,
    post_relatorio,
    resolve_tokens,
)
from extrator_arquivos import processar_dia
from ciclo_completo import ciclo_dia
from fechar_dia import fechar_dia
import db
import planos as planos_mod

app = Flask(__name__)
# SECRET_KEY assina o cookie de sessão. O fallback fixo estava COMMITADO — quem
# lesse o repo forjava sessão de admin. Em produção agora é obrigatória; fora de
# produção gera uma aleatória por processo (derruba as sessões a cada restart, o
# que é o comportamento correto pra dev).
_sk = os.environ.get("SECRET_KEY")
if not _sk:
    if os.environ.get("FLASK_ENV") == "production" or os.environ.get("RB_PRODUCAO") == "1":
        raise RuntimeError("SECRET_KEY não definida — obrigatória em produção.")
    import secrets as _secrets
    _sk = _secrets.token_urlsafe(48)
    logging.warning("SECRET_KEY ausente: usando chave aleatória (sessões caem no restart).")
app.secret_key = _sk
app.permanent_session_lifetime = timedelta(days=14)
# Explícito, não na sorte do default do navegador. Secure só fora de dev (em
# http://localhost o cookie com Secure é descartado e ninguém consegue logar).
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=(os.environ.get("RB_PRODUCAO") == "1"
                           or os.environ.get("FLASK_ENV") == "production"),
)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# ── Autenticação (usuário + senha, sessão) ────────────────────────────────────
# Rotas liberadas sem login (prefixos). Todo o resto exige sessão.
_PUBLICO = ("/login", "/logout", "/static", "/healthz", "/favicon",
            "/faturar/cron/rodar")  # protegido por token próprio (X-Trigger-Token)


# Cache curto da validação do usuário: {uid: (expira_em, dict_ou_None)}. Sem ele,
# revalidar a cada request bateria no banco em toda navegação.
_USER_CACHE: dict = {}
_USER_CACHE_S = 60
_user_cache_lock = threading.Lock()


def _usuario_valido(uid):
    """Usuário ainda existe e está ativo? (cache de 60s)"""
    agora = time.monotonic()
    with _user_cache_lock:
        hit = _USER_CACHE.get(uid)
        if hit and hit[0] > agora:
            return hit[1]
    try:
        u = db.get_usuario(uid)          # devolve None se inativo/inexistente
    except Exception:
        return {"id": uid}               # banco fora: não derruba quem já está logado
    with _user_cache_lock:
        _USER_CACHE[uid] = (agora + _USER_CACHE_S, u)
    return u


@app.before_request
def _exigir_login():
    p = request.path
    if any(p == x or p.startswith(x + "/") or p.startswith(x) for x in _PUBLICO):
        return None
    if session.get("uid"):
        # A sessão dura 14 dias e só era conferida no login: desativar um usuário
        # NÃO o desconectava, e uma mudança de papel só valia no próximo login.
        u = _usuario_valido(session["uid"])
        if u is None:
            session.clear()
            if p.startswith("/api/") or request.headers.get("Accept", "").startswith("application/json"):
                return jsonify({"error": "sessão encerrada"}), 401
            return redirect(url_for("login", next=p))
        if u.get("role"):
            session["role"] = u["role"]
        return None
    # API/JSON -> 401; navegador -> redireciona pro login
    if p.startswith("/api/") or request.headers.get("Accept", "").startswith("application/json"):
        return jsonify({"error": "não autenticado"}), 401
    return redirect(url_for("login", next=p))


@app.context_processor
def _injeta_usuario():
    return {"usuario_atual": {"username": session.get("username"),
                             "nome": session.get("nome"),
                             "role": session.get("role")} if session.get("uid") else None,
            "usuario_atual_id": session.get("uid"),
            "pendencias_abertas": db.contar_pendencias_front() if session.get("uid") else 0,
            "avisos_sem_guia": db.contar_avisos_nao_vistos() if session.get("uid") else 0}


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = db.autenticar(request.form.get("usuario", ""), request.form.get("senha", ""))
        if not u:
            return render_template("login.html", erro="Usuário ou senha inválidos."), 401
        session.permanent = True
        session["uid"] = u["id"]; session["username"] = u["username"]
        session["nome"] = u["nome"]; session["role"] = u["role"]
        dest = request.args.get("next") or "/"
        if not dest.startswith("/"):
            dest = "/"
        return redirect(dest)
    if session.get("uid"):
        return redirect("/")
    return render_template("login.html", erro=None)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/healthz")
def healthz():
    return "ok", 200


def _admin_ok():
    return session.get("role") == "admin"


@app.route("/usuarios", methods=["GET"])
def usuarios_page():
    if not _admin_ok():
        return ("Acesso restrito a administradores.", 403)
    return render_template("usuarios.html", usuarios=db.listar_usuarios())


@app.route("/usuarios/criar", methods=["POST"])
def usuarios_criar():
    if not _admin_ok():
        return jsonify({"error": "restrito a admin"}), 403
    try:
        u = db.criar_usuario(request.form.get("usuario", ""), request.form.get("senha", ""),
                             nome=request.form.get("nome", ""),
                             role=request.form.get("role", "user"))
        return jsonify({"ok": True, "usuario": u})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/usuarios/<int:uid>/senha", methods=["POST"])
def usuarios_senha(uid: int):
    if not _admin_ok():
        return jsonify({"error": "restrito a admin"}), 403
    try:
        db.resetar_senha(uid, request.form.get("senha", ""))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/usuarios/<int:uid>/ativo", methods=["POST"])
def usuarios_ativo(uid: int):
    if not _admin_ok():
        return jsonify({"error": "restrito a admin"}), 403
    if uid == session.get("uid"):
        return jsonify({"error": "não pode desativar a si mesmo"}), 400
    db.set_usuario_ativo(uid, request.form.get("ativo") == "1")
    with _user_cache_lock:                 # efeito imediato, sem esperar o cache
        _USER_CACHE.pop(uid, None)
    return jsonify({"ok": True})

# Escopo REDE UNNA — definido em config.py (evita import circular).
from config import CONVENIOS, SEGMENTOS, PLANOS, PLANOS_INATIVOS

# Inicializa o banco (cria tabelas se não existirem). Falha não derruba o app.
try:
    db.init_db()
    app.logger.info("Banco inicializado (%s).", db.DATABASE_URL.split("@")[-1])
    # Toda execução em 'running' no startup é zumbi (o processo que a iniciou
    # morreu) — marca como erro. SÓ sob gunicorn (produção): evita que um
    # `import app` local (teste) marque execuções reais do servidor.
    if "gunicorn" in sys.modules:
        _z = db.limpar_runs_travadas()
        if _z:
            app.logger.info("Limpas %d execução(ões) travada(s) no startup.", _z)
except Exception as _e:
    app.logger.error("Falha ao inicializar banco: %s", _e)

# ── Job store (em memória) ────────────────────────────────────────────────────
# Cresciam para sempre: nada nunca era removido daqui nem de _esteira_jobs, e o
# processo roda com --workers 1 --timeout 0 (não recicla). Cada job guarda o log
# completo e, no caso do /baixar_dia, o ZIP do dia INTEIRO em bytes. Com o tempo
# isso é OOM garantido — e o container reiniciando no meio de um faturamento.
_jobs: dict = {}
_jobs_lock = threading.Lock()
_JOB_TTL_S = 6 * 3600


def _purgar_jobs(store, lock=None, ttl=_JOB_TTL_S):
    """Remove jobs terminados há mais de `ttl`. Chamado ao criar um job novo —
    sem thread extra, sem custo quando o sistema está parado."""
    agora = time.monotonic()

    def _varre():
        for k, j in list(store.items()):
            if not (j.get("done") or j.get("status") in ("done", "error")):
                continue
            t_fim = j.get("t_fim")
            if t_fim is None:            # terminou sem carimbo: marca e deixa envelhecer
                j["t_fim"] = agora
                continue
            if (agora - t_fim) > ttl:
                store.pop(k, None)
    if lock:
        with lock:
            _varre()
    else:
        _varre()




def _run_job(job_id: str, data: str, convenios: list, segmentos: list) -> None:
    with _jobs_lock:
        _jobs[job_id]["status"] = "running"

    def progress(msg: str) -> None:
        with _jobs_lock:
            _jobs[job_id].setdefault("log", []).append(str(msg))

    try:
        zip_bytes, relatorio = processar_dia(data, convenios, segmentos, progress_cb=progress)
        # O ZIP do dia INTEIRO ficava na RAM até o restart do processo (o job store
        # nunca era limpo). Vai pro disco; o job guarda só o caminho.
        import tempfile
        fd, zpath = tempfile.mkstemp(prefix="_zipdia_", suffix=".zip")
        with os.fdopen(fd, "wb") as f:
            f.write(zip_bytes)
        del zip_bytes
        with _jobs_lock:
            _jobs[job_id].update(
                {"status": "done", "zip_path": zpath, "relatorio": relatorio}
            )
    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error("Erro no job %s:\n%s", job_id, tb)
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(exc), "traceback": tb})


def _run_ciclo_job(job_id: str, data: str, convenios: list, segmentos: list) -> None:
    with _jobs_lock:
        _jobs[job_id]["status"] = "running"

    def progress(msg: str) -> None:
        with _jobs_lock:
            _jobs[job_id].setdefault("log", []).append(str(msg))

    # /ciclo_dia ANEXA de verdade (ciclo_dia tem dry_run=False por padrão e a rota
    # não passa o parâmetro). É um SEGUNDO caminho de escrita, então precisa da
    # mesma trava da esteira — senão roda junto com /faturar/run ou com o cron no
    # mesmo dia e sobem 2x14 Chromium (o crash-loop conhecido). Reserva o dia
    # INTEIRO (todas as contas) porque o ciclo varre o dia, não uma unidade.
    _reservadas = []
    for _c in list(PLANOS) + [""]:
        if _esteira_reservar(data, _c, job_id):
            _reservadas.append(_c)
        else:
            for _r in _reservadas:
                _esteira_liberar(data, _r, job_id)
            msg = ("Já existe um faturamento em andamento para esse dia — "
                   "aguarde terminar antes de rodar o ciclo.")
            with _jobs_lock:
                _jobs[job_id].update({"status": "error", "error": msg})
            return
    try:
        relatorio = ciclo_dia(data, convenios, segmentos, progress_cb=progress)
        with _jobs_lock:
            _jobs[job_id].update({"status": "done", "relatorio": relatorio})
    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error("Erro no ciclo %s:\n%s", job_id, tb)
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(exc), "traceback": tb})
    finally:
        for _r in _reservadas:
            _esteira_liberar(data, _r, job_id)


def _run_fechar_job(job_id: str, data: str, dry_run: bool, plano: str = "odontoprev") -> None:
    """Job do 'Fechar dia' (orquestrador completo fechar_dia.py)."""
    with _jobs_lock:
        _jobs[job_id]["status"] = "running"

    def progress(msg: str) -> None:
        with _jobs_lock:
            _jobs[job_id].setdefault("log", []).append(str(msg))

    # Registra a execução no histórico (não bloqueia se o banco falhar).
    run_id = None
    try:
        run_id = db.criar_run(data, dry_run, plano=plano)
    except Exception as e:
        app.logger.error("Falha ao criar run no banco: %s", e)
    with _jobs_lock:
        _jobs[job_id]["run_id"] = run_id

    def _log_texto():
        with _jobs_lock:
            return "\n".join(_jobs[job_id].get("log", []))

    # TERCEIRO caminho que anexa (fechar_dia, pipeline antigo). Mesma trava dos
    # outros dois — em execução REAL não pode coincidir com /faturar/run, o cron
    # ou /ciclo_dia no mesmo dia.
    _reservadas = []
    if not dry_run:
        for _c in list(PLANOS) + [""]:
            if _esteira_reservar(data, _c, job_id):
                _reservadas.append(_c)
            else:
                for _r in _reservadas:
                    _esteira_liberar(data, _r, job_id)
                msg = ("Já existe um faturamento em andamento para esse dia — "
                       "aguarde terminar.")
                with _jobs_lock:
                    _jobs[job_id].update({"status": "error", "error": msg})
                return
    try:
        relatorio = fechar_dia(data, CONVENIOS, SEGMENTOS,
                               dry_run=dry_run, progress_cb=progress)
        if run_id is not None:
            try:
                db.finalizar_run_ok(run_id, relatorio, log_texto=_log_texto())
            except Exception as e:
                app.logger.error("Falha ao salvar run %s: %s", run_id, e)
        with _jobs_lock:
            _jobs[job_id].update({"status": "done", "relatorio": relatorio})
    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error("Erro no fechar_dia %s:\n%s", job_id, tb)
        if run_id is not None:
            try:
                db.finalizar_run_erro(run_id, str(exc) + "\n\n" + tb, log_texto=_log_texto())
            except Exception:
                pass
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(exc), "traceback": tb})
    finally:
        for _r in _reservadas:
            _esteira_liberar(data, _r, job_id)


def _run_glosa_job(job_id: str, dia: str, contas: list, checar: bool,
                   checar_demo: bool = True) -> None:
    """Job da atualização do panorama de glosas (3 unidades por padrão)."""
    with _jobs_lock:
        _jobs[job_id]["status"] = "running"

    def progress(msg: str) -> None:
        with _jobs_lock:
            _jobs[job_id].setdefault("log", []).append(str(msg))

    try:
        import time as _time
        from playwright.sync_api import sync_playwright
        from glosa_extrator import CONTAS, extrair_unidade
        lote = _time.strftime("%Y%m%d%H%M%S")
        alvo = [(u, l) for u, l in CONTAS if not contas or u in contas]
        total = 0
        with sync_playwright() as pw:
            for conta, label in alvo:
                progress(f"==== {label} ({conta}) — período até {dia} ====")
                r = extrair_unidade(pw, conta, label, dia, "_diag_glosa",
                                    checar_recursos=checar, checar_demonstrativo=checar_demo,
                                    log=progress)
                db.salvar_glosas(lote, dia, r["eventos"])
                total += len(r["eventos"])
                progress(f"[{label}] gravado ({len(r['eventos'])}).")
        try:
            rem = db.prune_glosa(lote, dia)
            if rem:
                progress(f"limpeza: {rem} evento(s) de lotes antigos do mês removidos.")
        except Exception as e:
            app.logger.error("prune glosa: %s", e)
        with _jobs_lock:
            _jobs[job_id].update({"status": "done", "lote": lote, "total": total})
    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error("Erro no glosa job %s:\n%s", job_id, tb)
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(exc), "traceback": tb})


def _run_desfecho_job(job_id: str, desde: str, contas: list) -> None:
    """Job da atualização do desfecho (status na RedeUna das guias que faturamos)."""
    with _jobs_lock:
        _jobs[job_id]["status"] = "running"

    def progress(msg: str) -> None:
        with _jobs_lock:
            _jobs[job_id].setdefault("log", []).append(str(msg))

    try:
        import time as _time
        from playwright.sync_api import sync_playwright
        from desfecho_extrator import extrair_desfechos_conta
        dia = _time.strftime("%d/%m/%Y")
        fat = db.guias_faturadas_por_nos(desde_dia=desde)
        fat = [f for f in fat if f.get("conta") and (not contas or f["conta"] in contas)]
        porconta = {}
        for f in fat:
            porconta.setdefault(f["conta"], {"unidade": f["unidade"], "guias": []})["guias"].append(f)
        total = sum(len(v["guias"]) for v in porconta.values())
        lote = _time.strftime("%Y%m%d%H%M%S")
        progress(f"{total} guia(s) em {len(porconta)} unidade(s) (faturadas desde {desde})")
        with sync_playwright() as pw:
            for conta, info in porconta.items():
                progress(f"==== {info['unidade']} ({conta}) — {len(info['guias'])} guia(s) ====")
                itens = extrair_desfechos_conta(pw, conta, info["unidade"], info["guias"],
                                                dia, log=progress)
                db.salvar_desfechos(lote, itens)
                progress(f"[{info['unidade']}] gravado ({len(itens)}).")
        with _jobs_lock:
            _jobs[job_id].update({"status": "done", "lote": lote, "total": total})
    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error("Erro no desfecho job %s:\n%s", job_id, tb)
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(exc), "traceback": tb})


def _run_anexacao_job(job_id: str, de: str, ate: str, contas: list, limite: int) -> None:
    """Job da varredura de anexação/faturamento (só-leitura, 3 unidades)."""
    with _jobs_lock:
        _jobs[job_id]["status"] = "running"

    def progress(msg: str) -> None:
        with _jobs_lock:
            _jobs[job_id].setdefault("log", []).append(str(msg))

    try:
        import time as _time
        from playwright.sync_api import sync_playwright
        from anexacao_extrator import CONTAS, varrer_unidade
        lote = _time.strftime("%Y%m%d%H%M%S")
        alvo = [(u, l) for u, l in CONTAS if not contas or u in contas]
        total = 0
        with sync_playwright() as pw:
            for conta, label in alvo:
                progress(f"==== {label} ({conta}) — {de} a {ate} ====")
                r = varrer_unidade(pw, conta, label, de, ate, limite=limite, log=progress)
                db.salvar_anexacao(lote, de, ate, r["gtos"])
                total += len(r["gtos"])
                progress(f"[{label}] gravado ({len(r['gtos'])}).")
        # varredura cumulativa do mês -> remove snapshots antigos do mesmo período
        try:
            rem = db.prune_anexacao(lote, de)
            if rem:
                progress(f"limpeza: {rem} registro(s) de varreduras antigas do período removidos.")
        except Exception as e:
            app.logger.error("prune anexacao: %s", e)
        with _jobs_lock:
            _jobs[job_id].update({"status": "done", "lote": lote, "total": total})
    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error("Erro no anexacao job %s:\n%s", job_id, tb)
        with _jobs_lock:
            _jobs[job_id].update({"status": "error", "error": str(exc), "traceback": tb})


def _glosa_atualizou_hoje() -> bool:
    """True se já há um lote de glosa capturado hoje (horário de Brasília)."""
    try:
        lotes = db.glosa_lotes(1)
        if not lotes or not lotes[0].get("captured_at"):
            return False
        d = datetime.fromisoformat(lotes[0]["captured_at"])
        if _TZ:
            if d.tzinfo is None:
                from datetime import timezone as _tzc
                d = d.replace(tzinfo=_tzc.utc)
            d = d.astimezone(_TZ)
            hoje = datetime.now(_TZ).date()
        else:
            hoje = datetime.now().date()
        return d.date() == hoje
    except Exception:
        return False


_glosa_ultima_tentativa = None  # data da última tentativa: garante 1x/dia (mesmo se falhar)


def _glosa_scheduler():
    """Atualiza o panorama de glosas 1x/dia (após GLOSA_UPDATE_HOUR, Brasília).
    gunicorn roda 1 worker -> sem concorrência de agendadores.
    Tenta no máximo 1x por dia — mesmo se o login falhar — pra não martelar o
    OdontoPrev de 30 em 30 min (o que dispara o rate-limit/bot-detection)."""
    global _glosa_ultima_tentativa
    try:
        hora = int(os.environ.get("GLOSA_UPDATE_HOUR", "6"))
    except ValueError:
        hora = 6
    while not _glosa_stop.is_set():
        try:
            # guarda periódico: execução presa em running há +3h = travada
            try:
                db.limpar_runs_travadas(horas=3)
            except Exception:
                pass
            agora = datetime.now(_TZ) if _TZ else datetime.now()
            if (agora.hour >= hora and not _glosa_atualizou_hoje()
                    and _glosa_ultima_tentativa != agora.date()):
                _glosa_ultima_tentativa = agora.date()  # marca a tentativa do dia (mesmo se falhar)
                dia = agora.strftime("%d/%m/%Y")
                jid = "auto" + uuid.uuid4().hex[:8]
                _purgar_jobs(_jobs, _jobs_lock)
                with _jobs_lock:
                    _jobs[jid] = {"status": "queued", "log": [], "kind": "glosa"}
                app.logger.info("Glosa auto-update iniciando (%s)…", dia)
                _run_glosa_job(jid, dia, [], True, True)
                app.logger.info("Glosa auto-update concluído.")
        except Exception as e:
            app.logger.error("Glosa scheduler: %s", e)
        _glosa_stop.wait(1800)  # re-checa a cada 30 min


_glosa_stop = threading.Event()
if os.environ.get("GLOSA_AUTO_UPDATE", "0") == "1":
    threading.Thread(target=_glosa_scheduler, daemon=True).start()


def _anexacao_atualizou_hoje() -> bool:
    try:
        lotes = db.anexacao_lotes(1)
        if not lotes or not lotes[0].get("captured_at"):
            return False
        d = datetime.fromisoformat(lotes[0]["captured_at"])
        if _TZ:
            if d.tzinfo is None:
                from datetime import timezone as _tzc
                d = d.replace(tzinfo=_tzc.utc)
            d = d.astimezone(_TZ)
            hoje = datetime.now(_TZ).date()
        else:
            hoje = datetime.now().date()
        return d.date() == hoje
    except Exception:
        return False


_anexacao_ultima_tentativa = None  # data da última tentativa: garante 1x/dia (mesmo se falhar)


def _anexacao_scheduler():
    """Varre anexação/faturamento das 3 unidades 1x/dia (após ANEXACAO_UPDATE_HOUR,
    default 7h Brasília — escalonado da glosa p/ não rodarem juntas).
    Tenta no máximo 1x por dia — mesmo se o login falhar — pra não martelar o
    OdontoPrev de 30 em 30 min (o que dispara o rate-limit/bot-detection)."""
    global _anexacao_ultima_tentativa
    try:
        hora = int(os.environ.get("ANEXACAO_UPDATE_HOUR", "7"))
    except ValueError:
        hora = 7
    while not _glosa_stop.is_set():
        try:
            agora = datetime.now(_TZ) if _TZ else datetime.now()
            if (agora.hour >= hora and not _anexacao_atualizou_hoje()
                    and _anexacao_ultima_tentativa != agora.date()):
                _anexacao_ultima_tentativa = agora.date()  # marca a tentativa do dia (mesmo se falhar)
                hoje = agora.strftime("%d/%m/%Y")
                de = "01/" + hoje[3:]
                jid = "anxauto" + uuid.uuid4().hex[:8]
                _purgar_jobs(_jobs, _jobs_lock)
                with _jobs_lock:
                    _jobs[jid] = {"status": "queued", "log": [], "kind": "anexacao"}
                app.logger.info("Anexação auto-update iniciando (%s a %s)…", de, hoje)
                _run_anexacao_job(jid, de, hoje, [], 0)
                app.logger.info("Anexação auto-update concluído.")
        except Exception as e:
            app.logger.error("Anexacao scheduler: %s", e)
        _glosa_stop.wait(1800)


if os.environ.get("ANEXACAO_AUTO_UPDATE", "0") == "1":
    threading.Thread(target=_anexacao_scheduler, daemon=True).start()


def _retry_scheduler():
    """LOOP DE RETRY do transitório (Fase 3): a cada ~20min re-tenta, DIRECIONADO, as
    guias que falharam por INFRA (gemini/rede/throttle/JWT) e estão devidas. Só o
    transitório entra na fila (classe_retry); externo/lógica NUNCA. Backoff
    exponencial + teto — depois de esgotar, vira pendência 'nossa, não recuperou'.
    Desligado por padrão (igual o faturar cron); ligue com RETRY_CRON=1."""
    while not _glosa_stop.is_set():
        try:
            import esteira
            res = esteira.processar_retries(
                gemini_key=os.environ.get("GEMINI_API_KEY"), k_attach=3,
                log=lambda m: app.logger.info("%s", m))
            if res.get("devidos"):
                app.logger.info("Retry loop: %s devida(s) em %s grupo(s).",
                                res.get("devidos"), res.get("grupos"))
        except Exception as e:
            app.logger.error("Retry scheduler: %s", e)
        # ~1 min entre ciclos: a 2a tentativa é IMEDIATA (backoff 0), então o poll
        # precisa ser curto pra "imediato" valer de verdade. Ciclo sem item devido é
        # só uma query barata (retries_devidos); anexa só quando há transitório na hora.
        _glosa_stop.wait(60)


if os.environ.get("RETRY_CRON", "0") == "1":
    threading.Thread(target=_retry_scheduler, daemon=True).start()


def _desfecho_atualizou_hoje() -> bool:
    try:
        lote = (db.desfecho_panorama() or {}).get("lote")
        if not lote:
            return False
        d = datetime.strptime(lote[:8], "%Y%m%d").date()
        hoje = datetime.now(_TZ).date() if _TZ else datetime.now().date()
        return d == hoje
    except Exception:
        return False


_desfecho_ultima_tentativa = None  # 1x/dia mesmo se falhar (não martelar o OdontoPrev)


def _desfecho_scheduler():
    """Atualiza o DESFECHO (status na RedeUna dos faturados do MÊS corrente) 1x/dia,
    após DESFECHO_UPDATE_HOUR (default 8h Brasília — escalonado da glosa/anexação).
    Tenta no máximo 1x/dia mesmo se falhar. Gated por DESFECHO_AUTO_UPDATE=1."""
    global _desfecho_ultima_tentativa
    try:
        hora = int(os.environ.get("DESFECHO_UPDATE_HOUR", "8"))
    except ValueError:
        hora = 8
    while not _glosa_stop.is_set():
        try:
            agora = datetime.now(_TZ) if _TZ else datetime.now()
            if (agora.hour >= hora and not _desfecho_atualizou_hoje()
                    and _desfecho_ultima_tentativa != agora.date()):
                _desfecho_ultima_tentativa = agora.date()
                # JANELA MÓVEL (default 120 dias = cobre todo o prazo de recurso, orto
                # inclusive). Assim a tela mantém o histórico pago/glosado/recorrer e não
                # encolhe pro mês corrente. Configurável por DESFECHO_DIAS.
                try:
                    _dias = int(os.environ.get("DESFECHO_DIAS", "120"))
                except ValueError:
                    _dias = 120
                desde = (agora.date() - timedelta(days=_dias)).strftime("%d/%m/%Y")
                jid = "desfauto" + uuid.uuid4().hex[:8]
                _purgar_jobs(_jobs, _jobs_lock)
                with _jobs_lock:
                    _jobs[jid] = {"status": "queued", "log": [], "kind": "desfecho"}
                app.logger.info("Desfecho auto-update iniciando (desde %s)…", desde)
                _run_desfecho_job(jid, desde, [])
                app.logger.info("Desfecho auto-update concluído.")
        except Exception as e:
            app.logger.error("Desfecho scheduler: %s", e)
        _glosa_stop.wait(1800)


if os.environ.get("DESFECHO_AUTO_UPDATE", "0") == "1":
    threading.Thread(target=_desfecho_scheduler, daemon=True).start()


# ── Faturamento automático diário (cron D-4 + reprocessa pendências) ────────────
def _faturar_rodou_hoje() -> bool:
    """Já rodou o faturamento automático hoje? (compara a última marca em Brasília)."""
    try:
        d = db.cron_faturar_last_at()
        if not d:
            return False
        if _TZ:
            if d.tzinfo is None:
                from datetime import timezone as _tzc
                d = d.replace(tzinfo=_tzc.utc)
            d = d.astimezone(_TZ)
            hoje = datetime.now(_TZ).date()
        else:
            hoje = datetime.now().date()
        return d.date() == hoje
    except Exception:
        return False


def _prazo_dias():
    try:
        return int(os.environ.get("FATURAR_PRAZO_DIAS", "7"))
    except ValueError:
        return 7


def _sla_dias_restantes(dia_str):
    """Dias que faltam pro prazo de faturamento estourar (dia do exame + prazo).
    None se a data não parseia. Negativo/0 = vencido."""
    from datetime import date
    d = db._parse_ddmmaaaa(dia_str)
    if not d:
        return None
    hoje = datetime.now(_TZ).date() if _TZ else date.today()
    return _prazo_dias() - (hoje - d).days


def _send_email(assunto, corpo_txt, corpo_html=None):
    """Envia email via SMTP (env). Retorna True/False. Pula se SMTP não configurado."""
    host = os.environ.get("SMTP_HOST")
    to = os.environ.get("ALERTA_EMAIL_TO")
    if not host or not to:
        app.logger.info("SLA email pulado: SMTP_HOST/ALERTA_EMAIL_TO não configurados.")
        return False
    try:
        import smtplib
        from email.message import EmailMessage
        port = int(os.environ.get("SMTP_PORT", "587"))
        user = os.environ.get("SMTP_USER")
        pwd = os.environ.get("SMTP_PASSWORD")
        remetente = os.environ.get("SMTP_FROM") or user or "radiobras@localhost"
        destinos = [e.strip() for e in to.split(",") if e.strip()]
        msg = EmailMessage()
        msg["Subject"] = assunto
        msg["From"] = remetente
        msg["To"] = ", ".join(destinos)
        msg.set_content(corpo_txt)
        if corpo_html:
            msg.add_alternative(corpo_html, subtype="html")
        with smtplib.SMTP(host, port, timeout=30) as s:
            s.ehlo()
            try:
                s.starttls(); s.ehlo()
            except Exception:
                pass
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)
        app.logger.info("SLA email enviado para %s", destinos)
        return True
    except Exception as e:
        app.logger.error("SLA email falhou: %s", str(e)[:140])
        return False


def _sla_status(sla):
    """(rótulo, cor) do estágio de urgência de um SLA. Assume sla <= 2."""
    if sla <= 0:
        return ("VENCIDA", "#7a0d0d")
    if sla == 1:
        return ("vence amanhã", "#b3261e")
    return ("faltam 2 dias", "#9a4d00")


def _enviar_alertas_sla():
    """Email diário das pendências dentro de 2 dias do prazo — vencidas, vence
    amanhã (1 dia) e faltam 2 dias. `sla <= 2` (e não `== 1`) dá margem de correção
    e é rede de segurança: se o cron pular um dia, a vencida ainda é alertada (e
    reforçada todo dia até resolver), em vez de o aviso de '1 dia' sumir pra sempre."""
    if os.environ.get("ALERTA_SLA", "1") == "0":
        return
    try:
        itens = db.listar_pendencias("abertas")
    except Exception:
        return
    urgentes = []
    for p in itens:
        s = _sla_dias_restantes(p.get("dia"))
        if s is not None and s <= 2:
            p["sla"] = s
            p["unidade"] = _plano_nome(p.get("conta")) or (p.get("conta") or "—")
            urgentes.append(p)
    if not urgentes:
        app.logger.info("SLA: nenhuma pendência dentro de 2 dias do prazo — sem email.")
        return
    urgentes.sort(key=lambda p: p["sla"])          # vencidas primeiro
    n_venc = sum(1 for p in urgentes if p["sla"] <= 0)
    n_d1 = sum(1 for p in urgentes if p["sla"] == 1)
    n_d2 = sum(1 for p in urgentes if p["sla"] == 2)
    partes = []
    if n_venc: partes.append(f"{n_venc} vencida(s)")
    if n_d1: partes.append(f"{n_d1} vence(m) amanhã")
    if n_d2: partes.append(f"{n_d2} em 2 dias")
    resumo = " · ".join(partes)

    linhas = "\n".join(
        f"  • [{_sla_status(p['sla'])[0]}] {p['unidade']} · dia {p['dia']} · GTO {p['gto']} "
        f"· {p.get('paciente') or '—'} — {p.get('motivo') or 'revisão'}" for p in urgentes)
    txt = (f"ATENÇÃO: {len(urgentes)} GTO(s) NÃO FATURADA(S) no limite do prazo "
           f"(prazo de {_prazo_dias()} dias da OdontoPrev): {resumo}.\n\n{linhas}\n\n"
           f"Resolva no PRORADIS ou faça a correção na origem hoje, senão o prazo estoura.\n"
           f"Painel de pendências: /pendencias")
    rows = "".join(
        f"<tr><td style='padding:6px 10px;color:{_sla_status(p['sla'])[1]};font-weight:700'>"
        f"{_sla_status(p['sla'])[0]}</td>"
        f"<td style='padding:6px 10px'>{p['unidade']}</td>"
        f"<td style='padding:6px 10px'>{p['dia']}</td>"
        f"<td style='padding:6px 10px'><b>{p['gto']}</b></td>"
        f"<td style='padding:6px 10px'>{(p.get('paciente') or '—')}</td>"
        f"<td style='padding:6px 10px'>{(p.get('motivo') or 'revisão')}</td></tr>"
        for p in urgentes)
    html = (f"<div style='font-family:Arial,sans-serif'>"
            f"<h2 style='color:#b3261e'>⚠️ {len(urgentes)} GTO(s) no limite do prazo</h2>"
            f"<p>{resumo}. Prazo de <b>{_prazo_dias()} dias</b> da OdontoPrev. Resolva hoje.</p>"
            f"<table style='border-collapse:collapse;font-size:13px' border='1'>"
            f"<tr style='background:#f0f0f0'><th style='padding:6px 10px'>Prazo</th>"
            f"<th style='padding:6px 10px'>Unidade</th>"
            f"<th style='padding:6px 10px'>Dia</th><th style='padding:6px 10px'>GTO</th>"
            f"<th style='padding:6px 10px'>Paciente</th><th style='padding:6px 10px'>Motivo</th></tr>"
            f"{rows}</table></div>")
    assunto = (f"⚠️ RadioBras — {resumo} (prazo de faturamento)")
    _send_email(assunto, txt, html)


_faturar_cron_running = threading.Event()


def _faturar_cron_rodar():
    """Roda D-4 nas 3 unidades + reprocessa os dias com pendência aberta dentro do
    prazo. Faturamento REAL (anexa). Idempotente: já faturado é pulado; pendência
    resolvida no PRORADIS é faturada agora e fecha sozinha."""
    if _faturar_cron_running.is_set():
        app.logger.info("Cron faturar já em execução — pulando disparo.")
        return
    _faturar_cron_running.set()
    try:
        _faturar_cron_body()
    finally:
        _faturar_cron_running.clear()


# Tempo máximo que uma reserva vale sem ser liberada. É rede de segurança para o
# caso do `finally` não rodar (container morto no meio da execução) — uma esteira
# real leva minutos, não horas.
_ESTEIRA_TTL_S = 2 * 3600


def _esteira_reservar(dia, conta, tag, ttl=_ESTEIRA_TTL_S):
    """Reserva (dia, conta) para UMA esteira. Devolve a tag se conseguiu, None se
    já há outra rodando. Duas esteiras no mesmo dia/unidade sobem 2x15 Chromium
    (crash-loop do container), podem anexar em DUPLICIDADE — a idempotência do
    upload lê os anexos ANTES de enviar, então duas leituras simultâneas concluem
    as duas que "falta anexar" — e disputam o mesmo login no portal.

    A reserva é AUTOCONTIDA: guarda o próprio instante. A versão anterior perguntava
    a `_esteira_jobs` se o dono ainda estava vivo, mas só o /faturar/run registra a
    tag lá — o cron usa 'cron-<conta>-<dia>' e o /fechar usa um job_id que vive em
    `_jobs`. Para esses dois, `job` era None, o `if` não barrava e a reserva alheia
    era SOBRESCRITA: o cenário real (cron às 5h + alguém clicando em Faturar) passava
    direto, que era exatamente o que a trava deveria impedir."""
    chave = (dia, conta or "_")
    agora = time.monotonic()
    with _esteira_ativas_lock:
        atual = _esteira_ativas.get(chave)
        if atual and (agora - atual["t"]) < ttl:
            return None
        _esteira_ativas[chave] = {"tag": tag, "t": agora}
        return tag


def _esteira_liberar(dia, conta, tag):
    chave = (dia, conta or "_")
    with _esteira_ativas_lock:
        if (_esteira_ativas.get(chave) or {}).get("tag") == tag:
            _esteira_ativas.pop(chave, None)


def _dia_alvo_cron(hoje):
    """Dia-alvo do cron diário: D-4 (quatro dias atrás), formatado 'DD/MM/AAAA'."""
    from datetime import timedelta
    return (hoje - timedelta(days=4)).strftime("%d/%m/%Y")


def _faturar_cron_body():
    from datetime import date
    try:
        prazo = int(os.environ.get("FATURAR_PRAZO_DIAS", "7"))
    except ValueError:
        prazo = 7
    hoje = datetime.now(_TZ).date() if _TZ else date.today()
    target = _dia_alvo_cron(hoje)
    combos = {(c, target) for c in PLANOS}                       # D-4 nas 3 unidades
    combos |= set(db.dias_com_pendencia_aberta(prazo))           # + pendências no prazo
    gkey = os.environ.get("GEMINI_API_KEY")
    from esteira import rodar_esteira
    ndias = nfat = 0
    for conta, dia in sorted(combos):
        _tag = f"cron-{conta}-{dia}"
        if not _esteira_reservar(dia, conta, _tag):
            app.logger.warning("Cron faturar %s %s PULADO: já há execução em andamento",
                               conta, dia)
            continue
        _logs = []          # captura o log da execução p/ persistir (igual ao web)
        try:
            resumo = rodar_esteira(dia, 6, 3, 5, log=_logs.append,
                                   gemini_key=gkey, k_attach=3, dry_run=False,
                                   conta=conta, senha_portal=db.get_portal_senha(conta))
            # SEMPRE salva (antes só salvava com pendentes>0): quando a última
            # pendência do dia era resolvida, o cron pulava a gravação e ela NUNCA
            # fechava — o alerta de SLA seguia cobrando guia já faturada.
            # A gravação tem try/except próprio (igual ao web): um hiccup do banco
            # DEPOIS de a guia já ter faturado não pode virar "FALHOU" — a guia já
            # está anexada; o próximo run vê 'ja_anexada' e não duplica.
            if resumo:
                try:
                    db.salvar_execucao(resumo, _logs)
                except Exception as e:
                    app.logger.error("Cron faturar %s %s: anexou mas falhou ao "
                                     "gravar: %s", conta, dia, str(e)[:120])
            nfat += (resumo or {}).get("anexado_ok", 0) or 0
            ndias += 1
            app.logger.info("Cron faturar %s %s: fat=%s pend=%s", conta, dia,
                            (resumo or {}).get("anexado_ok"), (resumo or {}).get("pendentes"))
        except Exception as e:
            app.logger.error("Cron faturar %s %s FALHOU: %s", conta, dia, str(e)[:120])
            try:
                db.salvar_execucao_falha(dia, conta, False, str(e), _logs)
            except Exception:
                pass
        finally:
            _esteira_liberar(dia, conta, _tag)
    db.cron_marcar_faturar(target)
    app.logger.info("Cron faturar concluído: %s execução(ões), %s faturada(s).", ndias, nfat)
    try:
        _enviar_alertas_sla()      # email só dos que vencem amanhã (1 dia p/ o prazo)
    except Exception as e:
        app.logger.error("Alerta SLA falhou: %s", str(e)[:120])


def _faturar_scheduler():
    """Dispara o faturamento automático 1x/dia (após FATURAR_CRON_HOUR, Brasília).
    Desligado por padrão — ligue com FATURAR_CRON=1. gunicorn 1 worker -> sem
    concorrência de agendadores."""
    try:
        hora = int(os.environ.get("FATURAR_CRON_HOUR", "5"))
    except ValueError:
        hora = 5
    while not _glosa_stop.is_set():
        try:
            agora = datetime.now(_TZ) if _TZ else datetime.now()
            if agora.hour >= hora and not _faturar_rodou_hoje():
                app.logger.info("Cron faturar iniciando…")
                _faturar_cron_rodar()
        except Exception as e:
            app.logger.error("Faturar scheduler: %s", e)
        _glosa_stop.wait(1800)  # re-checa a cada 30 min


if os.environ.get("FATURAR_CRON", "0") == "1":
    threading.Thread(target=_faturar_scheduler, daemon=True).start()


# ── Resumo de faturamentos por email (automático, 1x/dia) ───────────────────────
def _resumo_fat_enviado_hoje() -> bool:
    try:
        d = db.cron_resumo_fat_last_at()
        if not d:
            return False
        if _TZ:
            if d.tzinfo is None:
                from datetime import timezone as _tzc
                d = d.replace(tzinfo=_tzc.utc)
            d = d.astimezone(_TZ)
            hoje = datetime.now(_TZ).date()
        else:
            hoje = datetime.now().date()
        return d.date() == hoje
    except Exception:
        return False


_resumo_fat_lock = threading.Lock()


def _resumo_fat_tentar(ignorar_hora: bool = False):
    """Envia o resumo dos faturamentos da semana no máximo 1x/dia (dedupe via
    banco). Thread-safe. Sem `ignorar_hora`, só envia após RESUMO_FAT_HOUR."""
    with _resumo_fat_lock:
        if _resumo_fat_enviado_hoje():
            return
        if not ignorar_hora:
            try:
                hora = int(os.environ.get("RESUMO_FAT_HOUR", "8"))
            except ValueError:
                hora = 8
            agora = datetime.now(_TZ) if _TZ else datetime.now()
            if agora.hour < hora:
                return
        _enviar_resumo_faturamentos()
        db.cron_marcar_resumo_fat()


def _resumo_fat_scheduler():
    """Envia o resumo de faturamentos 1x/dia (após RESUMO_FAT_HOUR, Brasília,
    default 8h). Sem faturamentos reais na semana ou sem SMTP, nada é enviado
    (mas o dia é marcado, pra não retentar em loop). 1 worker -> sem corrida."""
    while not _glosa_stop.is_set():
        try:
            _resumo_fat_tentar()
        except Exception as e:
            app.logger.error("Resumo fat scheduler: %s", e)
        _glosa_stop.wait(1800)


if os.environ.get("RESUMO_FAT_AUTO", "1") != "0":
    threading.Thread(target=_resumo_fat_scheduler, daemon=True).start()
    # garante o envio de hoje logo após o boot, independente da hora (1x/dia)
    threading.Timer(60, lambda: _resumo_fat_tentar(ignorar_hora=True)).start()


def _trigger_token_ok() -> bool:
    """Auth por token pra acionamento via API/automação (sem sessão de navegador).
    Header 'X-Trigger-Token' == env FATURAR_TRIGGER_TOKEN. Comparação constant-time.
    Só vale se o token estiver configurado (env não-vazio)."""
    import hmac
    tok = (os.environ.get("FATURAR_TRIGGER_TOKEN") or "").strip()
    hdr = (request.headers.get("X-Trigger-Token") or "").strip()
    return bool(tok) and hmac.compare_digest(tok, hdr)


@app.route("/faturar/cron/rodar", methods=["POST"])
def faturar_cron_rodar_now():
    """Dispara o faturamento automático sob demanda — roda em background.
    Autoriza por ADMIN logado OU por token de API (X-Trigger-Token)."""
    if not (_admin_ok() or _trigger_token_ok()):
        return jsonify({"error": "apenas admin ou token válido"}), 403
    if _faturar_cron_running.is_set():
        return jsonify({"error": "cron já está rodando"}), 409
    threading.Thread(target=_faturar_cron_rodar, daemon=True).start()
    return jsonify({"ok": True, "msg": "faturamento automático disparado"})


@app.route("/faturar/cron/status")
def faturar_cron_status():
    last = db.cron_faturar_last_at()
    return jsonify({"rodando": _faturar_cron_running.is_set(),
                    "ligado": os.environ.get("FATURAR_CRON", "0") == "1",
                    "ultima": last.isoformat() if last else None})


def _email_resumo_semana():
    """(assunto, txt, html) do resumo do que foi implementado na semana.
    Usado pelo teste de email — confirma o SMTP entregando conteúdo útil."""
    itens = [
        ("Faturamento automático diário (05/07)",
         "O robô roda sozinho todo dia (~5h): fatura o dia D-4 nas 3 unidades da Rede "
         "Una e reprocessa pendências ainda dentro do prazo. Idempotente (não duplica "
         "anexo) e já ligado em produção. Prazo da OdontoPrev confirmado em 7 dias."),
        ("Alerta de prazo / SLA (05/07, ampliado 09/07)",
         "Painel /pendencias mostra cada pendência com selo de urgência (vencida / vence "
         "amanhã / 2 / 3 dias). O e-mail de alerta agora cobre vencidas + vence amanhã + "
         "faltam 2 dias (antes só 'amanhã'), ordenado por urgência — reduz o risco de "
         "perder um prazo."),
        ("Backlog de revisão humana (03/07)",
         "Toda GTO não faturada vira uma tarefa com checkbox no painel /pendencias, "
         "guardada no banco durável (Supabase)."),
        ("Login do portal mais seguro (03/07)",
         "Login que falha aborta com erro claro (fim do 'sucesso' silencioso); a senha "
         "passou a ser por código de conta na tela /portal."),
        ("Diagnóstico do sistema (09/07)",
         "A tela /api/diag passou a mostrar se o robô rodou hoje, o tamanho do backlog "
         "e as últimas execuções — auditoria num relance."),
    ]
    assunto = "RadioBras — Resumo da semana (03–10/07): faturamento automático, alerta de prazo e diagnóstico"
    intro = ("Resumo do que foi implementado no sistema RadioBras Digital nesta semana "
             "(03 a 10/07/2026):")
    txt = intro + "\n\n" + "\n\n".join(
        f"{i}. {t}\n   {d}" for i, (t, d) in enumerate(itens, 1))
    txt += ("\n\n— E-mail enviado pelo próprio sistema para validar a configuração de "
            "envio (Hostinger).")
    blocos = "".join(
        f"<li style='margin:0 0 12px'><b>{t}</b><br>"
        f"<span style='color:#333'>{d}</span></li>" for t, d in itens)
    html = (f"<div style='font-family:Arial,sans-serif;max-width:640px'>"
            f"<h2 style='color:#0b6b4f'>RadioBras — Resumo da semana</h2>"
            f"<p>{intro}</p>"
            f"<ol style='font-size:14px;line-height:1.5;padding-left:18px'>{blocos}</ol>"
            f"<p style='color:#888;font-size:12px'>E-mail enviado pelo próprio sistema "
            f"para validar a configuração de envio (Hostinger).</p></div>")
    return assunto, txt, html


def _whatsapp_ok() -> bool:
    try:
        import notificador
        return notificador.whatsapp_configurado()
    except Exception:
        return False


@app.route("/alerta/testar-whatsapp", methods=["POST"])
def alerta_testar_whatsapp():
    """Manda uma mensagem de teste pro dono — confirma instância, token e número.
    Sem isso o canal só seria testado no dia em que a rodada quebrasse."""
    if not _admin_ok():
        return jsonify({"error": "apenas admin"}), 403
    import notificador
    if not notificador.whatsapp_configurado():
        return jsonify({"ok": False,
                        "msg": "Falta UAZAPI_TOKEN e/ou ALERTA_WHATSAPP_TO no ambiente."})
    ok = notificador.enviar_whatsapp(
        "✅ *RadioBras* — canal de alerta técnico ligado." + chr(10) * 2 +
        "É por aqui que você vai receber falha de sistema: rodada que aborta, guia "
        "que falhou por problema nosso e retry que não recuperou. "
        "A operação não vê nada disso no painel.")
    return jsonify({"ok": ok, "msg": "Enviado — confira o WhatsApp." if ok
                    else "Não enviou. Confira UAZAPI_HOST/UAZAPI_TOKEN e se a "
                         "instância está conectada."})


@app.route("/alerta/testar-email", methods=["POST"])
def alerta_testar_email():
    """Envia um email de teste (admin) — confirma o SMTP com o resumo da semana."""
    if not _admin_ok():
        return jsonify({"error": "apenas admin"}), 403
    assunto, txt, html = _email_resumo_semana()
    ok = _send_email(assunto, txt, html)
    return jsonify({"ok": ok,
                    "msg": "Email enviado — confira a caixa de entrada." if ok
                    else "Não enviou. Confira SMTP_HOST/PORT/USER/PASSWORD e ALERTA_EMAIL_TO."})


def _email_resumo_faturamentos(dias: int = 7):
    """(assunto, txt, html) com os faturamentos REAIS dos últimos `dias`
    (execuções não-dry-run do banco). Retorna (None, None, None) se não houver."""
    hoje = datetime.now(_TZ).date() if _TZ else datetime.now().date()
    limite = hoje - timedelta(days=dias)
    execs = []
    for e in db.listar_execucoes(200):
        if e.get("dry_run"):
            continue
        c = e.get("criado_em")
        if not c:
            continue
        cd = c.astimezone(_TZ).date() if (_TZ and c.tzinfo) else c.date()
        if cd < limite:
            continue
        e["quando"] = cd
        e["unidade"] = _plano_nome(e.get("conta")) or (e.get("conta") or "—")
        execs.append(e)
    if not execs:
        return None, None, None
    execs.sort(key=lambda e: (e["quando"], e["unidade"]))
    tot_fat = sum(e.get("faturadas", 0) or 0 for e in execs)
    tot_pend = sum(e.get("pendentes", 0) or 0 for e in execs)
    assunto = (f"RadioBras — Faturamentos da semana: {tot_fat} GTO(s) faturada(s) "
               f"em {len(execs)} execução(ões)")
    intro = (f"Resumo dos faturamentos reais dos últimos {dias} dias "
             f"({limite.strftime('%d/%m')} a {hoje.strftime('%d/%m/%Y')}):")
    total = (f"TOTAL: {tot_fat} GTO(s) faturada(s) · {tot_pend} pendência(s) "
             f"· {len(execs)} execução(ões).")
    linhas = "\n".join(
        f"  • {e['quando'].strftime('%d/%m')} · {e['unidade']} · dia do exame {e.get('dia') or '—'} "
        f"— {e.get('faturadas', 0)} faturada(s), {e.get('pendentes', 0)} pendente(s)"
        for e in execs)
    txt = f"{intro}\n\n{total}\n\n{linhas}\n\nPainel: /relatorios · Pendências: /pendencias"
    rows = "".join(
        f"<tr><td style='padding:6px 10px'>{e['quando'].strftime('%d/%m')}</td>"
        f"<td style='padding:6px 10px'>{e['unidade']}</td>"
        f"<td style='padding:6px 10px'>{e.get('dia') or '—'}</td>"
        f"<td style='padding:6px 10px;text-align:center'><b>{e.get('faturadas', 0)}</b></td>"
        f"<td style='padding:6px 10px;text-align:center'>{e.get('pendentes', 0)}</td></tr>"
        for e in execs)
    html = (f"<div style='font-family:Arial,sans-serif;max-width:680px'>"
            f"<h2 style='color:#0b6b4f'>Faturamentos da semana</h2>"
            f"<p>{intro}</p>"
            f"<p style='font-size:15px'><b>{total}</b></p>"
            f"<table style='border-collapse:collapse;font-size:13px' border='1'>"
            f"<tr style='background:#f0f0f0'><th style='padding:6px 10px'>Rodou em</th>"
            f"<th style='padding:6px 10px'>Unidade</th><th style='padding:6px 10px'>Dia do exame</th>"
            f"<th style='padding:6px 10px'>Faturadas</th><th style='padding:6px 10px'>Pendentes</th></tr>"
            f"{rows}</table></div>")
    return assunto, txt, html


def _enviar_resumo_faturamentos():
    """Monta e envia o resumo de faturamentos da semana. Loga o resultado."""
    assunto, txt, html = _email_resumo_faturamentos()
    if not assunto:
        app.logger.info("Resumo faturamentos: nenhuma execução real na semana — sem email.")
        return
    ok = _send_email(assunto, txt, html)
    app.logger.info("Resumo faturamentos enviado=%s", ok)


@app.route("/alerta/resumo-faturamentos", methods=["POST"])
def alerta_resumo_faturamentos():
    """Agenda (ou envia já) o resumo dos faturamentos da semana por email (admin).
    ?delay=<segundos> agenda pra daqui a N segundos (default 300 = 5 min)."""
    if not _admin_ok():
        return jsonify({"error": "apenas admin"}), 403
    try:
        delay = int(request.args.get("delay", "300"))
    except ValueError:
        delay = 300
    delay = max(0, min(delay, 3600))
    if delay == 0:
        _enviar_resumo_faturamentos()
        return jsonify({"ok": True, "msg": "Resumo enviado agora — confira a caixa."})
    threading.Timer(delay, _enviar_resumo_faturamentos).start()
    return jsonify({"ok": True, "agendado_s": delay,
                    "msg": f"Agendado — o resumo chega em ~{delay // 60} min."})


def _rotulo_sla(sla):
    """(rótulo, cor) de qualquer SLA (inclusive no-prazo e sem-data)."""
    if sla is None:
        return ("sem data", "#666")
    if sla <= 0:
        return ("VENCIDA", "#7a0d0d")
    if sla == 1:
        return ("vence amanhã", "#b3261e")
    if sla == 2:
        return ("faltam 2 dias", "#9a4d00")
    return (f"no prazo ({sla}d)", "#0b6b4f")


def _email_pendencias_abertas():
    """(assunto, txt, html) com TODAS as pendências abertas agora, ordenadas por
    urgência (vencidas primeiro). (None, None, None) se não houver."""
    try:
        itens = db.listar_pendencias("abertas")
    except Exception:
        return None, None, None
    if not itens:
        return None, None, None
    for p in itens:
        p["sla"] = _sla_dias_restantes(p.get("dia"))
        p["unidade"] = _plano_nome(p.get("conta")) or (p.get("conta") or "—")
    itens.sort(key=lambda p: (p["sla"] is None, p["sla"] if p["sla"] is not None else 9999))
    total = len(itens)
    n_venc = sum(1 for p in itens if p["sla"] is not None and p["sla"] <= 0)
    assunto = f"RadioBras — {total} pendência(s) aberta(s) · {n_venc} vencida(s)"
    linhas = "\n".join(
        f"  • [{_rotulo_sla(p['sla'])[0]}] {p['unidade']} · dia {p.get('dia') or '—'} · GTO {p.get('gto')} "
        f"· {p.get('paciente') or '—'} — {p.get('motivo') or 'revisão'}" for p in itens)
    txt = (f"Todas as pendências abertas agora: {total} ({n_venc} vencidas). "
           f"Prazo OdontoPrev {_prazo_dias()} dias.\n\n{linhas}\n\nPainel: /pendencias")
    rows = "".join(
        f"<tr><td style='padding:5px 9px;color:{_rotulo_sla(p['sla'])[1]};font-weight:700'>{_rotulo_sla(p['sla'])[0]}</td>"
        f"<td style='padding:5px 9px'>{p['unidade']}</td>"
        f"<td style='padding:5px 9px'>{p.get('dia') or '—'}</td>"
        f"<td style='padding:5px 9px'><b>{p.get('gto')}</b></td>"
        f"<td style='padding:5px 9px'>{p.get('paciente') or '—'}</td>"
        f"<td style='padding:5px 9px'>{p.get('motivo') or 'revisão'}</td></tr>" for p in itens)
    html = (f"<div style='font-family:Arial,sans-serif'>"
            f"<h2 style='color:#b3261e'>{total} pendência(s) aberta(s) · {n_venc} vencida(s)</h2>"
            f"<p>Prazo de faturamento OdontoPrev: <b>{_prazo_dias()} dias</b>. Painel: /pendencias</p>"
            f"<table style='border-collapse:collapse;font-size:12px' border='1'>"
            f"<tr style='background:#f0f0f0'><th style='padding:5px 9px'>Prazo</th>"
            f"<th style='padding:5px 9px'>Unidade</th><th style='padding:5px 9px'>Dia</th>"
            f"<th style='padding:5px 9px'>GTO</th><th style='padding:5px 9px'>Paciente</th>"
            f"<th style='padding:5px 9px'>Motivo</th></tr>{rows}</table></div>")
    return assunto, txt, html


@app.route("/alerta/pendencias-agora", methods=["POST"])
def alerta_pendencias_agora():
    """Envia AGORA um email com TODAS as pendências abertas (admin)."""
    if not _admin_ok():
        return jsonify({"error": "apenas admin"}), 403
    assunto, txt, html = _email_pendencias_abertas()
    if not assunto:
        return jsonify({"ok": False, "msg": "Nenhuma pendência aberta no momento."})
    ok = _send_email(assunto, txt, html)
    return jsonify({"ok": ok, "msg": "Email enviado — confira a caixa." if ok
                    else "Não enviou (confira SMTP_HOST/ALERTA_EMAIL_TO)."})


@app.route("/api/pendencias")
def api_pendencias():
    """Relação única de pendências (todas as contas juntas), c/ unidade e SLA.
    ?status=abertas|resolvidas|todas (default abertas). Admin."""
    if not _admin_ok():
        return jsonify({"error": "apenas admin"}), 403
    status = request.args.get("status", "abertas")
    itens = db.listar_pendencias(status)
    for p in itens:
        p["sla"] = _sla_dias_restantes(p.get("dia"))
        p["unidade"] = _plano_nome(p.get("conta")) or (p.get("conta") or "—")
        p["prazo_rotulo"] = _rotulo_sla(p["sla"])[0]
    itens.sort(key=lambda p: (p["sla"] is None, p["sla"] if p["sla"] is not None else 9999))
    return jsonify({"total": len(itens), "prazo_dias": _prazo_dias(), "pendencias": itens})


# ── Rotas ─────────────────────────────────────────────────────────────────────

@app.route("/")
def home():
    """Tela principal — Dashboard (com 'Executar Agora')."""
    return render_template("dashboard.html")


@app.route("/fechar-simples")
def fechar_simples():
    """Tela enxuta de 'Fechar o dia' (fallback)."""
    return render_template("fechar.html")


@app.route("/gtos")
def gtos_page():
    """Detalhe de GTOs / funil de um dia (mockup 2)."""
    return render_template("gtos.html")


# ── Pendências (worklist por urgência) ─────────────────────────────────────────
# Substitui a antiga "Revisão humana": mesma fila, mas agrupada por bucket de SLA
# (do mais urgente ao menos) e com a coluna "quem age" (responsável por item).
_BUCKETS_SLA = [
    ("venc", "Vencidas"),
    ("d1", "Vence amanhã"),
    ("d2", "Em 2 dias"),
    ("d3", "Em 3 dias"),
    ("no_prazo", "No prazo"),
]


def _sla_bucket(sla):
    """Bucket de urgência a partir do SLA (dias restantes). Sem data cai em 'no_prazo'."""
    if sla is None:
        return "no_prazo"
    if sla <= 0:
        return "venc"
    if sla == 1:
        return "d1"
    if sla == 2:
        return "d2"
    if sla == 3:
        return "d3"
    return "no_prazo"


def _filtra_por_dia(itens, de_iso, ate_iso):
    """Recorta a lista pelo dia do EXAME (campo 'dia', DD/MM/AAAA). Datas em ISO,
    como o <input type=date> manda.

    Com o prazo de faturamento correndo, a pergunta do dia a dia e "o que vence
    primeiro" — e isso se responde recortando o periodo.

    FALHA ABERTA de proposito: data invalida devolve a lista INTEIRA. Um filtro que
    quebra e some com tudo faz a operacao achar que a fila zerou, e pendencia que
    ninguem ve e pendencia que vence."""
    import datetime as _dt

    def _iso(v):
        try:
            return _dt.datetime.strptime(str(v or "").strip(), "%Y-%m-%d").date()
        except Exception:
            return None
    de, ate = _iso(de_iso), _iso(ate_iso)
    if not de and not ate:
        return itens
    out = []
    for p in itens:
        d = db._parse_ddmmaaaa(p.get("dia"))
        if not d:
            continue          # com periodo pedido, guia sem data nao entra por omissao
        if de and d < de:
            continue
        if ate and d > ate:
            continue
        out.append(p)
    return out


@app.route("/pendencias")
def pendencias_page():
    """Worklist: itens não faturados que precisam de ação humana, por urgência."""
    import datetime as _dt
    status = request.args.get("status", "abertas")
    if status not in ("abertas", "resolvidas", "todas"):
        status = "abertas"
    itens = db.listar_pendencias(status=status)
    import arquivos_pendencia as _ap
    _ap_base = _ap.base_dir()
    _gtos = [p.get("gto") for p in itens]
    _leituras = db.leituras_por_gtos(_gtos)   # o que a IA leu (evidência)
    _tent = db.tentativas_por_gtos(_gtos)      # p/ saber se um transitório já esgotou
    for p in itens:
        p["unidade"] = _plano_nome(p.get("conta")) or (p.get("conta") or "—")
        p["sla"] = _sla_dias_restantes(p.get("dia"))   # dias p/ o prazo (None/negativo=vencido)
        p["bucket"] = _sla_bucket(p["sla"])
        _chave, _quem, _acao = db.classificar_pendencia(p.get("motivo"), p.get("categoria") or "")
        p["acao"] = _acao
        p["chave"] = _chave   # p/ a tela distinguir 'nome não bate' de 'ilegível'
        # título CURTO do motivo (headline) — a linha fica enxuta; o texto completo
        # abre ao clicar. Ex.: "Paciente não encontrado no PRORADIS".
        p["titulo"] = getattr(db, "_TITULO_GRUPO", {}).get(_chave) or "Pendência"
        # TIPO — regra do dono: NOSSO (transitório em reprocessamento) NÃO é do front;
        # DELES é 'aguardar' (terceiro) ou 'conferir' (humano). Esgotado deixou de ser
        # nosso → vira 'conferir/investigar' no front.
        _classe = db.classe_efetiva(p.get("motivo") or "", p.get("categoria") or "",
                                    _tent.get(str(p.get("gto")), 0))
        p["classe"] = _classe
        # 22/08: o gate passou a ser "e NOSSO?", nao "esta em reprocessamento?". Uma
        # falha nossa que ESGOTOU o retry continua sendo nossa — antes voltava pro
        # operador como "Investigar" e ele nao tinha o que fazer com bug nosso.
        if db.eh_nosso(p.get("motivo") or "", p.get("categoria") or ""):
            p["tipo"] = "interno"
            p["responsavel"] = "Investigar" if _classe == "esgotado" else "Reprocessar"
        elif _classe == "externo":
            p["tipo"] = "aguardar"; p["responsavel"] = _quem
        else:                                   # conferencia (olho humano no doc)
            p["tipo"] = "conferir"; p["responsavel"] = _quem
        _l = _leituras.get(str(p.get("gto"))) or {}
        p["exames_gto"] = _l.get("exames_gto")
        p["exames_lidos"] = _l.get("exames_lidos")
        p["lido"] = _l.get("lido")
        # o nome LIDO no documento, ao lado do botao de confirmar: o clique e
        # irreversivel e a operadora precisa ver o que esta escrito no papel.
        p["nomes_lidos"] = db.nomes_lidos_resumo(_l.get("lido"))
        # quantos arquivos guardados desta guia — 0 quando o volume nao esta
        # montado ou a execucao e antiga. So mostra o bloco se houver algo; link
        # que abre em 404 e pior do que nao ter link.
        try:
            p["n_arquivos"] = len(_ap.listar(_ap_base, p.get("conta"), p.get("dia"),
                                             p.get("gto"), p.get("paciente")))                 if _ap_base else 0
        except Exception:
            p["n_arquivos"] = 0
    # front (o usuário vê) tira TUDO que é INTERNO (nosso) — em reprocessamento ou já
    # esgotado. O resto aparece agrupado por dia, nada escondido da operação. Falha
    # nossa não some em silêncio: vai pro WhatsApp do dono (notificador.py) e pro loop
    # de retry. A urgência (vencida/no prazo) fica marcada no cabeçalho de cada dia.
    front = [p for p in itens if p["tipo"] != "interno"]
    # FILTRO DE DATA (23/08, pedido do dono). Vem depois do corte do interno: o que
    # e falha nossa nao aparece nesta tela com filtro nenhum.
    _de = (request.args.get("de") or "").strip()
    _ate = (request.args.get("ate") or "").strip()
    front = _filtra_por_dia(front, _de, _ate)
    sla_ct = {"venc": 0, "d1": 0, "d2": 0, "d3": 0}
    for p in front:
        if p.get("resolvido"):
            continue
        s = p.get("sla")
        if s is None:
            continue
        if s <= 0:      sla_ct["venc"] += 1
        elif s == 1:    sla_ct["d1"] += 1
        elif s == 2:    sla_ct["d2"] += 1
        elif s == 3:    sla_ct["d3"] += 1
    _ordem_bucket = {"venc": 0, "d1": 1, "d2": 2, "d3": 3, "no_prazo": 4}
    _dias_sem = ["segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo"]
    por_dia = {}
    for p in front:
        por_dia.setdefault(p.get("dia") or "—", []).append(p)
    grupos = []
    for dia, lst in por_dia.items():
        lst.sort(key=lambda p: (_ordem_bucket.get(p["bucket"], 9),
                                (p.get("unidade") or "").lower(), p.get("gto") or ""))
        _d = db._parse_ddmmaaaa(dia)
        titulo = (f"{dia} · {_dias_sem[_d.weekday()]}" if _d else (dia or "sem data"))
        pior = min((p["bucket"] for p in lst), key=lambda b: _ordem_bucket.get(b, 9))
        grupos.append({"chave": pior, "titulo": titulo, "dia": dia, "itens": lst,
                       "abertas": sum(1 for p in lst if not p.get("resolvido")),
                       "total": len(lst)})
    grupos.sort(key=lambda g: db._parse_ddmmaaaa(g["dia"]) or _dt.date.max)
    return render_template("pendencias.html", grupos=grupos, grupos_venc=None,
                           itens=front, status=status, prazo=_prazo_dias(), sla_ct=sla_ct,
                           n_abertas=sum(1 for p in front if not p.get("resolvido")),
                           n_vencidas=0, de=_de, ate=_ate)


# ── ARQUIVOS DA PENDENCIA (22/08, pedido da Andrea) ─────────────────────────
# "criar pasta com imagens resolvidas para casos de nao conseguir ler solicitacoes,
# depois ela anexa tudo". Os arquivos ja foram baixados na rodada; a esteira copia
# pra /dados/pendencias/<plano>/<data>/<GTO_PACIENTE>/ antes de limpar.
#
# NADA e servido como estatico: sao laudo e imagem de paciente (LGPD). Estas rotas
# passam pelo guard de login como o resto do app, e o caminho e resolvido a partir
# do ID da pendencia no BANCO — se viesse pela URL, daria pra pedir a pasta de
# qualquer outro paciente.
def _pasta_da_pendencia(pid):
    """(base, pendencia) ou (None, None) se o recurso estiver desligado/inexistente."""
    import arquivos_pendencia as ap
    base = ap.base_dir()
    if not base:
        return None, None
    pend = db.pendencia_por_id(pid)
    return (base, pend) if pend else (None, None)


@app.route("/pendencias/<int:pid>/arquivos")
def pendencia_arquivos(pid):
    """O que ha na pasta desta guia — para a tela montar miniatura e download."""
    import arquivos_pendencia as ap
    base, pend = _pasta_da_pendencia(pid)
    if not base:
        return jsonify({"itens": [], "motivo": "pasta nao configurada"})
    itens = ap.listar(base, pend["conta"], pend["dia"], pend["gto"], pend["paciente"])
    return jsonify({"itens": itens, "gto": pend["gto"],
                    "paciente": pend["paciente"], "dia": pend["dia"]})


@app.route("/pendencias/<int:pid>/arquivo/<path:nome>")
def pendencia_arquivo(pid, nome):
    """Um arquivo. `inline` para ver no navegador; sem ele, baixa."""
    import arquivos_pendencia as ap
    base, pend = _pasta_da_pendencia(pid)
    if not base:
        return ("Pasta de arquivos não configurada.", 404)
    caminho = ap.caminho_do_arquivo(base, pend["conta"], pend["dia"], pend["gto"],
                                    pend["paciente"], nome)
    if not caminho:
        # nome suspeito (../) ou arquivo inexistente — a mesma resposta para os dois,
        # para nao contar a quem sonda se o arquivo existe
        return ("Arquivo não encontrado.", 404)
    # XSS COM O COOKIE DA OPERADORA: os arquivos vem do prontuario do PRORADIS, ou
    # seja de TERCEIRO. Servir .html/.svg inline renderiza a pagina dentro da sessao
    # dela. So imagem e PDF abrem inline; o resto baixa. Mais o nosniff, para o
    # navegador nao adivinhar tipo pelo conteudo.
    _ext = os.path.splitext(caminho)[1].lower()
    _pode_inline = _ext in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf")
    inline = request.args.get("inline") == "1" and _pode_inline
    resp = send_file(caminho, as_attachment=not inline,
                     download_name=os.path.basename(caminho))
    resp.headers["X-Content-Type-Options"] = "nosniff"
    return resp


@app.route("/pendencias/<int:pid>/arquivos.zip")
def pendencia_arquivos_zip(pid):
    """Tudo de uma vez — e o que a operadora quer quando vai anexar no portal."""
    import zipfile
    import arquivos_pendencia as ap
    base, pend = _pasta_da_pendencia(pid)
    if not base:
        return ("Pasta de arquivos não configurada.", 404)
    itens = ap.listar(base, pend["conta"], pend["dia"], pend["gto"], pend["paciente"])
    if not itens:
        return ("Nenhum arquivo guardado para esta guia.", 404)
    bio = io.BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
        for it in itens:
            caminho = ap.caminho_do_arquivo(base, pend["conta"], pend["dia"],
                                            pend["gto"], pend["paciente"], it["nome"])
            if caminho:
                z.write(caminho, arcname=it["nome"])
    bio.seek(0)
    _pac = re.sub(r"[^A-Za-z0-9]+", "_", str(pend.get("paciente") or "")).strip("_")
    nome_zip = f"{pend['gto']}_{_pac}.zip" if _pac else f"{pend['gto']}.zip"
    return send_file(bio, as_attachment=True, download_name=nome_zip,
                     mimetype="application/zip")


@app.route("/tecnico")
def tecnico_page():
    """FILA TÉCNICA — a tela do DONO, separada da tela da operação.

    Regra dele (22/08): falha de sistema não aparece pro pessoal da RadioBras. O
    `/pendencias` e o `/relatorios/pendencias` são deles; esta é dele. Os alertas do
    WhatsApp apontam para cá — antes apontavam para a tela da operação, e ele caía no
    meio das pendências da Andrea tendo que caçar a seção técnica (feedback 23/08)."""
    if not _admin_ok():
        return ("Acesso restrito — esta tela é da fila técnica.", 403)
    import notificador
    itens = db.listar_pendencias(status="abertas")
    uniq = {}
    for p in itens:
        uniq[(p.get("conta"), p.get("dia"), p.get("gto"))] = p
    nossas = [p for p in uniq.values()
              if db.eh_nosso(p.get("motivo") or "", p.get("categoria") or "")]
    _tent = db.tentativas_por_gtos([p.get("gto") for p in nossas])
    grupos = {}
    for p in nossas:
        p["unidade"] = _plano_nome(p.get("conta"))
        p["causa"] = notificador._resumir_causa(p.get("motivo"))
        p["tentativas"] = _tent.get(str(p.get("gto")), 0)
        p["esgotada"] = db.classe_efetiva(p.get("motivo") or "",
                                          p.get("categoria") or "",
                                          p["tentativas"]) == "esgotado"
        grupos.setdefault(p["causa"], []).append(p)
    lista = sorted(grupos.items(), key=lambda kv: -len(kv[1]))
    for _causa, gs in lista:
        gs.sort(key=lambda x: (x.get("dia") or "", x.get("gto") or ""))
    return render_template("tecnico.html", grupos=lista, total=len(nossas),
                           esgotadas=sum(1 for p in nossas if p["esgotada"]),
                           pausado=db.retry_pausado(),
                           pausa=db.retry_pausa_info())


@app.route("/pendencias/<int:pid>/resolver", methods=["POST"])
def pendencias_resolver(pid):
    obs = (request.form.get("obs") or (request.json.get("obs") if request.is_json else None)) if (request.form or request.is_json) else None
    db.resolver_pendencia(pid, session.get("username") or session.get("nome") or "?", obs=obs)
    if request.is_json or request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "abertas": db.contar_pendencias_front()})
    return redirect(url_for("pendencias_page", status=request.args.get("status", "abertas")))


@app.route("/pendencias/<int:pid>/reabrir", methods=["POST"])
def pendencias_reabrir(pid):
    db.reabrir_pendencia(pid)
    if request.is_json or request.headers.get("X-Requested-With") == "fetch":
        return jsonify({"ok": True, "abertas": db.contar_pendencias_front()})
    return redirect(url_for("pendencias_page", status=request.args.get("status", "abertas")))


@app.route("/pendencias/<int:pid>/confirmar", methods=["POST"])
def pendencias_confirmar(pid):
    """SINAL VERDE HUMANO (feature 13/08): o usuário abriu a pendência (ilegível /
    nome não bate), conferiu que a solicitação É do paciente e libera o faturamento.
    Registra a confirmação (a responsabilidade é dele) e dispara um faturamento REAL
    só desta guia — a decisão libera a trava do nome/cobertura; o LAUDO segue
    obrigatório. POST (anexa no portal, irreversível)."""
    if not session.get("uid"):
        return jsonify({"error": "faça login"}), 403
    with db.SessionLocal() as s:
        p = s.get(db.Pendencia, pid)
        if not p:
            return jsonify({"error": "pendência não encontrada"}), 404
        gto, conta, dia = str(p.gto), p.conta, p.dia
    quem = session.get("username") or session.get("nome") or "?"
    db.confirmar_nome(gto, conta, dia, quem)   # o sinal verde fica gravado
    # dispara faturamento REAL só desta guia (background). Se o portal não estiver
    # acessível agora (proxy), a confirmação fica gravada e o próximo processamento
    # fatura. Idempotente: se já anexado, não duplica.
    jid = uuid.uuid4().hex[:8]
    if not _esteira_reservar(dia, conta, jid):
        return jsonify({"ok": True, "confirmado": True,
                        "msg": "Confirmado ✔ — já há uma execução em andamento nessa "
                               "unidade; vai faturar nela."})
    gkey = os.environ.get("GEMINI_API_KEY")
    senha = db.get_portal_senha(conta or None)

    def _go():
        try:
            from esteira import rodar_esteira
            r = rodar_esteira(dia, m_download=3, n_desc=3, k_leitura=5, log=None,
                              gemini_key=gkey, k_attach=3, dry_run=False,
                              conta=(conta or None), senha_portal=senha,
                              apenas_gtos=[gto])
            if r:
                db.salvar_execucao(r, None)
        except Exception:
            pass
        finally:
            _esteira_liberar(dia, conta, jid)

    threading.Thread(target=_go, daemon=True).start()
    return jsonify({"ok": True, "confirmado": True,
                    "msg": "Confirmado ✔ — faturando esta guia em segundo plano "
                           "(atualize em ~1 min). Se ainda faltar o laudo, fatura "
                           "quando o laudo sair."})


# ── Avisos "exame sem guia" (laudo pronto sem GTO) — aviso, não pendência ──────
@app.route("/api/avisos")
def api_avisos():
    """Avisos NÃO vistos, para o modal + banner. Qualquer usuário logado."""
    if not session.get("uid"):
        return jsonify({"total": 0, "avisos": []})
    itens = db.listar_avisos("nao_vistos")
    for a in itens:
        a["unidade"] = _plano_nome(a.get("conta")) or (a.get("conta") or "—")
    return jsonify({"total": len(itens), "avisos": itens})


@app.route("/avisos/<int:aid>/ciente", methods=["POST"])
def aviso_ciente(aid):
    """Marca UM aviso como visto (Ciente). Grava quem/quando."""
    if not session.get("uid"):
        return jsonify({"ok": False}), 403
    n = db.marcar_aviso_visto(aid, session.get("username") or session.get("nome") or "?")
    return jsonify({"ok": True, "restantes": n})


@app.route("/avisos/ciente-todos", methods=["POST"])
def avisos_ciente_todos():
    """Marca TODOS os avisos abertos como vistos."""
    if not session.get("uid"):
        return jsonify({"ok": False}), 403
    db.marcar_todos_avisos_vistos(session.get("username") or session.get("nome") or "?")
    return jsonify({"ok": True, "restantes": 0})


# ── Senha do portal RedeUna/OdontoPrev (por código) ────────────────────────────
def _testar_login_portal(conta, senha):
    """Teste de login ao vivo no OdontoPrev (user = código da conta). (ok, msg)."""
    try:
        from playwright.sync_api import sync_playwright
        from extrator_odontoprev import login_odonto
        with sync_playwright() as pw:
            br, ctx, pg = login_odonto(pw, conta, senha)
            url = pg.url
            br.close()
        return True, f"Login OK ({url[:50]})"
    except Exception as e:
        return False, str(e)[:160]


@app.route("/portal")
def portal_page():
    """Cadastro da senha do portal por código (plano)."""
    status = db.listar_portal_status()
    planos = [{"conta": c, "nome": _plano_nome(c), "label": v.get("label", c),
               "tem": status.get(c, {}).get("tem", False),
               "atualizado_em": status.get(c, {}).get("atualizado_em"),
               "por": status.get(c, {}).get("por")}
              for c, v in PLANOS.items()]
    return render_template("portal.html", planos=planos)


@app.route("/portal/senha", methods=["POST"])
def portal_salvar():
    conta = (request.form.get("conta") or "").strip()
    senha = request.form.get("senha") or ""
    if conta not in PLANOS:
        return jsonify({"error": "código inválido"}), 400
    if not senha:
        return jsonify({"error": "senha vazia"}), 400
    try:
        db.set_portal_senha(conta, senha, username=session.get("username"))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)[:120]}), 500


@app.route("/portal/testar", methods=["POST"])
def portal_testar():
    conta = (request.form.get("conta") or "").strip()
    senha = request.form.get("senha") or db.get_portal_senha(conta)
    if conta not in PLANOS:
        return jsonify({"error": "código inválido"}), 400
    if not senha:
        return jsonify({"error": "sem senha para testar"}), 400
    ok, msg = _testar_login_portal(conta, senha)
    return jsonify({"ok": ok, "msg": msg})


@app.route("/relatorio")
def index():
    """Tela antiga (relatório analítico xlsx + download ZIP)."""
    return render_template("index.html", convenios=CONVENIOS, segmentos=SEGMENTOS)


# ── Relatório de execução (visual + PDF) ───────────────────────────────────────
# Cada status vira um grupo visual, com rótulo, cor e o "porquê" determinístico.
STATUS_META = {
    "ENVIADO":     {"label": "Anexado",            "cls": "ok",   "icon": "✓",
                    "desc": "Laudo e imagens anexados na GTO."},
    "PRONTO":      {"label": "Pronto para anexar", "cls": "ok",   "icon": "✓",
                    "desc": "Arquivos encontrados — seriam anexados numa execução real (simulação)."},
    "JA_ANEXADO":  {"label": "Já estava anexado",  "cls": "ok",   "icon": "✓",
                    "desc": "Os arquivos já estavam na GTO (nada a reenviar)."},
    "SEM_LAUDO":   {"label": "Sem laudo",          "cls": "warn", "icon": "!",
                    "desc": "Exame ainda não laudado no PRORADIS."},
    "SEM_IMAGENS": {"label": "Sem imagens",        "cls": "warn", "icon": "!",
                    "desc": "Sem imagens disponíveis no PRORADIS."},
    "SEM_MATCH":   {"label": "Não localizado",     "cls": "bad",  "icon": "✕",
                    "desc": "Paciente não localizado no PRORADIS."},
    "AMBIGUO":     {"label": "Ambíguo",            "cls": "bad",  "icon": "✕",
                    "desc": "Mais de um paciente com o mesmo nome — precisa conferência."},
    "ERRO_UPLOAD": {"label": "Erro ao anexar",     "cls": "bad",  "icon": "✕",
                    "desc": "Falha no envio do anexo."},
}
_ORDEM_STATUS = ["ENVIADO", "PRONTO", "JA_ANEXADO", "SEM_LAUDO", "SEM_IMAGENS",
                 "SEM_MATCH", "AMBIGUO", "ERRO_UPLOAD"]


def _fmt_run_datas(run: dict) -> dict:
    """Adiciona início/fim formatados (Brasília) e duração ao dict da run."""
    def _parse(s):
        if not s:
            return None
        try:
            dt = datetime.fromisoformat(s)
            return dt.astimezone(_TZ) if _TZ else dt
        except Exception:
            return None
    ini, fim = _parse(run.get("started_at")), _parse(run.get("finished_at"))
    run["ini_fmt"] = ini.strftime("%d/%m/%Y %H:%M") if ini else "—"
    run["fim_fmt"] = fim.strftime("%d/%m/%Y %H:%M") if fim else "—"
    if ini and fim:
        seg = max(int((fim - ini).total_seconds()), 0)
        m, s = divmod(seg, 60)
        run["dur_fmt"] = (f"{m}m {s}s" if m else f"{s}s")
    else:
        run["dur_fmt"] = "—"
    return run


def _agrupar_run(run: dict) -> list:
    """Agrupa os itens por status (na ordem de _ORDEM_STATUS), pulando vazios."""
    itens = run.get("itens", []) or []
    por_status = {}
    for it in itens:
        st = (it.get("status") or "?").upper()
        por_status.setdefault(st, []).append(it)
    grupos = []
    vistos = set()
    for st in _ORDEM_STATUS + sorted(por_status.keys()):
        if st in vistos or st not in por_status:
            continue
        vistos.add(st)
        meta = STATUS_META.get(st, {"label": st.title(), "cls": "warn", "icon": "•", "desc": ""})
        grupos.append({"status": st, "meta": meta, "itens": por_status[st]})
    return grupos


@app.route("/relatorio/run/<int:run_id>")
def relatorio_run(run_id: int):
    """Relatório visual de uma execução (o que foi feito, o que não, e por quê)."""
    run = db.run_detalhe(run_id)
    if not run:
        return ("Execução não encontrada.", 404)
    _fmt_run_datas(run)
    embed = request.args.get("embed") in ("1", "true", "yes")
    return render_template("relatorio_run.html", run=run,
                           grupos=_agrupar_run(run), pdf=False, embed=embed)


@app.route("/relatorio/run/<int:run_id>.pdf")
def relatorio_run_pdf(run_id: int):
    """Mesmo relatório, renderizado em PDF pelo Chromium (Playwright). Download 1-clique."""
    run = db.run_detalhe(run_id)
    if not run:
        return ("Execução não encontrada.", 404)
    _fmt_run_datas(run)
    html = render_template("relatorio_run.html", run=run,
                           grupos=_agrupar_run(run), pdf=True)
    from playwright.sync_api import sync_playwright
    try:
        with sync_playwright() as pw:
            br = pw.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            pg = br.new_page()
            pg.set_content(html, wait_until="networkidle")
            pdf_bytes = pg.pdf(format="A4", print_background=True,
                               margin={"top": "12mm", "bottom": "12mm",
                                       "left": "10mm", "right": "10mm"})
            br.close()
    except Exception as exc:
        app.logger.error("Falha ao gerar PDF da run %s: %s", run_id, exc)
        return (f"Falha ao gerar PDF: {exc}", 500)
    nome = f"relatorio_{(run.get('dia') or '').replace('/', '-')}_run{run_id}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name=nome)


# ── Árvore de decisão (documento vivo: do botão ao fim, fiel ao código) ────────
@app.route("/arvore-decisao")
def arvore_decisao():
    """Documento que descreve, em linguagem humana, TODAS as decisões do sistema
    de faturamento — do clique em 'Faturar' até o fim. Pedido do dono p/ auditar
    se o sistema decide certo. Reflete o código em produção."""
    return render_template("arvore_decisao.html", pdf=False)


@app.route("/arvore-decisao.pdf")
def arvore_decisao_pdf():
    """A mesma árvore, em PDF P&B (Chromium/Playwright). Download 1-clique."""
    html = render_template("arvore_decisao.html", pdf=True)
    from playwright.sync_api import sync_playwright
    try:
        with sync_playwright() as pw:
            br = pw.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            pg = br.new_page()
            pg.set_content(html, wait_until="networkidle")
            pdf_bytes = pg.pdf(format="A4", print_background=True,
                               margin={"top": "12mm", "bottom": "12mm",
                                       "left": "10mm", "right": "10mm"})
            br.close()
    except Exception as exc:
        app.logger.error("Falha ao gerar PDF da árvore de decisão: %s", exc)
        return (f"Falha ao gerar PDF: {exc}", 500)
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name="arvore_decisao_faturamento.pdf")


# ── Relatório de execuções (pipeline novo) — dentro de /relatorios ────────────
def _fmt_quando(dt):
    """Datetime UTC -> 'DD/MM/AAAA HH:MM' no horário de Brasília (UTC-3)."""
    if not dt:
        return ""
    try:
        from datetime import timedelta
        return (dt - timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(dt)


def _plano_nome(conta):
    """Nome amigável do plano a partir do código da conta."""
    if not conta:
        return "—"
    p = PLANOS.get(conta)
    return p["label"] if p else conta


@app.route("/relatorios/execucoes")
def relatorios_execucoes():
    execs = db.listar_execucoes(300)
    for e in execs:
        e["quando"] = _fmt_quando(e.get("criado_em"))
        e["plano"] = _plano_nome(e.get("conta"))
    tot = {
        "n": len(execs),
        "faturadas": sum(e["faturadas"] for e in execs),
        "nao_faturadas": sum(e["nao_faturadas"] for e in execs),
        "tempo_total": sum(e["tempo_total"] for e in execs),
        "tempo_medio": round(sum(e["tempo_total"] for e in execs) / len(execs)) if execs else 0,
    }
    return render_template("execucoes.html", execs=execs, tot=tot)


@app.route("/relatorios/execucao/<int:eid>")
def relatorios_execucao(eid: int):
    ex = db.get_execucao(eid)
    if not ex:
        return ("Execução não encontrada.", 404)
    ex["quando"] = _fmt_quando(ex.get("criado_em"))
    ex["plano"] = _plano_nome(ex.get("conta"))
    return render_template("execucao.html", ex=ex, pdf=False)


@app.route("/relatorios/execucao/<int:eid>/log")
def relatorios_execucao_log(eid: int):
    """Log técnico BRUTO da execução — a evidência de cada decisão, guia por guia.

    Existia gravado em Execucao.log desde sempre e não tinha por onde ler: a
    pergunta "por que essa guia não faturou?" só se respondia por dedução. Aqui
    está o que o robô viu — nomes lidos, arquivos baixados, o que entrou e saiu
    do plano de anexação, e por quê."""
    ex = db.get_execucao(eid)
    if not ex:
        return ("Execução não encontrada.", 404)
    cab = (f"# execução {eid} · dia {ex.get('dia')} · conta {ex.get('conta')}"
           f"{' · SIMULAÇÃO' if ex.get('dry_run') else ''}\n"
           f"# {ex.get('faturadas')} faturadas · {ex.get('nao_faturadas')} não faturadas\n")
    if ex.get("erro"):
        cab += f"# ABORTOU: {ex['erro']}\n"
    corpo = ex.get("log") or "(esta execução não gravou log)"
    return Response(cab + "\n" + corpo, mimetype="text/plain; charset=utf-8")


@app.route("/relatorios/execucao/<int:eid>.pdf")
def relatorios_execucao_pdf(eid: int):
    ex = db.get_execucao(eid)
    if not ex:
        return ("Execução não encontrada.", 404)
    ex["quando"] = _fmt_quando(ex.get("criado_em"))
    ex["plano"] = _plano_nome(ex.get("conta"))
    html = render_template("execucao.html", ex=ex, pdf=True)
    from playwright.sync_api import sync_playwright
    try:
        with sync_playwright() as pw:
            br = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            pg = br.new_page()
            pg.set_content(html, wait_until="networkidle")
            pdf_bytes = pg.pdf(format="A4", print_background=True,
                               margin={"top": "12mm", "bottom": "12mm", "left": "10mm", "right": "10mm"})
            br.close()
    except Exception as exc:
        return (f"Falha ao gerar PDF: {exc}", 500)
    nome = f"execucao_{(ex.get('dia') or '').replace('/', '-')}_{eid}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name=nome)


@app.route("/relatorios/execucao/<int:eid>.xlsx")
def relatorios_execucao_xlsx(eid: int):
    ex = db.get_execucao(eid)
    if not ex:
        return ("Execução não encontrada.", 404)
    import pandas as pd

    def _anexado(it):
        base = "laudo + imagens"
        return base + (f" + solicitação ({it['solicitacao']})" if it.get("solicitacao") else "")
    df_f = pd.DataFrame([{"GTO": i["gto"], "Paciente": i["paciente"],
                          "Exames (GTO)": i.get("exames_gto") or "",
                          "O que foi anexado": _anexado(i)}
                         for i in ex["faturadas_itens"]])
    df_n = pd.DataFrame([{"GTO": i["gto"], "Paciente": i["paciente"],
                          "Exames (GTO)": i.get("exames_gto") or "",
                          "Motivo (não faturada)": i.get("motivo") or ""}
                         for i in ex["nao_faturadas_itens"]])
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    # colunas largas (com quebra de texto) — pra não ficar "embolado"
    LARGAS = {"O que foi anexado", "Motivo (não faturada)"}

    def _estilizar(ws, df):
        head_fill = PatternFill("solid", fgColor="0F7A4F")
        head_font = Font(bold=True, color="FFFFFF", size=11)
        borda = Border(bottom=Side(style="thin", color="DDE5E0"))
        ws.row_dimensions[1].height = 26
        for ci, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left", wrap_text=False)
            # largura: maior conteúdo da coluna, com tetos por tipo
            larga = col in LARGAS
            vals = [str(col)] + [str(v) for v in df[col].tolist() if v is not None]
            maxlen = max((len(v) for v in vals), default=10)
            ws.column_dimensions[cell.column_letter].width = min(maxlen + 3, 70 if larga else 32)
        # corpo: quebra de texto nas colunas largas + borda leve + alinhamento topo
        for ri in range(2, ws.max_row + 1):
            for ci, col in enumerate(df.columns, start=1):
                c = ws.cell(row=ri, column=ci)
                c.alignment = Alignment(vertical="top", wrap_text=(col in LARGAS))
                c.border = borda
                c.font = Font(size=10)
        ws.freeze_panes = "A2"          # cabeçalho fixo ao rolar

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        fat = df_f if not df_f.empty else pd.DataFrame([{"(sem faturadas)": "—"}])
        nao = df_n if not df_n.empty else pd.DataFrame([{"(sem não faturadas)": "—"}])
        fat.to_excel(w, sheet_name="Faturadas", index=False)
        nao.to_excel(w, sheet_name="Não faturadas", index=False)
        _estilizar(w.sheets["Faturadas"], fat)
        _estilizar(w.sheets["Não faturadas"], nao)
    buf.seek(0)
    nome = f"execucao_{(ex.get('dia') or '').replace('/', '-')}_{eid}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/relatorios/execucao/<int:eid>.json")
def relatorios_execucao_json(eid: int):
    ex = db.get_execucao(eid)
    if not ex:
        return jsonify({"error": "não encontrada"}), 404
    ex.pop("criado_em", None)  # datetime não serializa
    return jsonify(ex)


# ── Faturar dia (UI do pipeline novo) ─────────────────────────────────────────
@app.route("/faturar")
def faturar_page():
    planos = [{"codigo": c, "label": p["label"]} for c, p in PLANOS.items()]
    return render_template("faturar.html", planos=planos, planos_inativos=PLANOS_INATIVOS)


# Execuções em andamento por (dia, conta) — impede 2 esteiras na MESMA GTO, que
# sobem 2x14 Chromium e já derrubaram o container (crash-loop).
_esteira_ativas = {}
_esteira_ativas_lock = threading.Lock()


@app.route("/faturar/run", methods=["POST"])
def faturar_run():
    """Dispara a esteira. POST — nunca GET: isto ANEXA documento no portal do
    convênio, então não pode ser acionável por link, imagem ou prefetch."""
    import time as _time
    data = (request.form.get("data") or "").strip()
    if not data:
        return jsonify({"error": "informe a data"}), 400
    plano = (request.form.get("plano") or "").strip()
    if plano and plano not in PLANOS:
        return jsonify({"error": "plano inválido"}), 400
    # padrão = DRY. Faturamento real exige dry=0 explícito.
    dry = (request.form.get("dry") or "1") != "0"

    jid = uuid.uuid4().hex[:8]
    if not _esteira_reservar(data, plano, jid):
        return jsonify({"error": "Já existe uma execução em andamento para esse "
                                 "dia e unidade."}), 409

    review_dir = f"/tmp/esteira_rev/{jid}"
    job = {"log": [], "done": False, "resumo": None, "error": None,
           "review_dir": review_dir, "execucao_id": None, "t0": _time.monotonic(),
           # conta e dry ficam no job para o /api/diag conseguir dizer O QUE está
           # rodando, não só que há algo rodando
           "dia": data, "conta": plano, "dry": dry}
    _purgar_jobs(_esteira_jobs)
    _esteira_jobs[jid] = job
    gkey = os.environ.get("GEMINI_API_KEY")

    senha_portal = db.get_portal_senha(plano or None)

    def _go():
        try:
            from esteira import rodar_esteira
            job["resumo"] = rodar_esteira(data, 6, 3, 5, lambda m: job["log"].append(m),
                                          gemini_key=gkey, review_dir=review_dir,
                                          k_attach=3, dry_run=dry, conta=(plano or None),
                                          senha_portal=senha_portal)
            # GRAVA SEMPRE, inclusive DRY. Antes a simulação não deixava rastro:
            # quando a operadora relatava um problema, não havia execução nem log
            # para consultar. O que continua valendo só para execução REAL é a
            # criação de pendência (dentro de salvar_execucao).
            if job["resumo"]:
                try:
                    job["execucao_id"] = db.salvar_execucao(job["resumo"], job["log"])
                except Exception as e:
                    job["log"].append(f"(falha ao salvar: {str(e)[:80]})")
            # 'resultados' carrega o objeto inteiro de cada GTO (inclusive as
            # leituras cruas do Gemini) e ninguém lê depois daqui — o que a tela usa
            # é 'decisoes', e o durável já foi pro banco. Segurava dezenas de MB por
            # execução, para sempre.
            if job["resumo"]:
                job["resumo"].pop("resultados", None)
        except Exception as e:
            job["error"] = str(e)
            job["log"].append(f"ERRO: {str(e)[:140]}")
            # REGISTRA A FALHA. Antes a execucao que dava erro sumia: nao virava
            # faturamento, nem pendencia, nem historico. A operadora dizia "eu
            # rodei esse dia" e nao havia como confirmar nem descobrir por que
            # nada aconteceu (caso 21/07 Camacari).
            try:
                job["execucao_id"] = db.salvar_execucao_falha(
                    data, plano, dry, str(e), job["log"])
            except Exception as e2:
                job["log"].append(f"(falha ao registrar o erro: {str(e2)[:80]})")
        finally:
            job["done"] = True
            _esteira_liberar(data, plano, jid)

    threading.Thread(target=_go, daemon=True).start()
    return jsonify({"job": jid})


def _esteira_progress(job: dict) -> dict:
    """Deriva progresso amigável (%, mensagem, ETA em seg) a partir do log do job."""
    import time as _time
    import re as _re
    log = job.get("log", [])
    pend = sum(1 for l in log if ">>> PENDENTE" in l)
    baix = sum(1 for l in log if "-> BAIXADO" in l)
    anex = sum(1 for l in log if "[ANEX" in l)
    # tempos (s desde o início) das DECISÕES concluídas — base estável p/ ETA
    dec_t = []
    for l in log:
        if "[DEC" in l and "mem=" in l:
            m = _re.match(r"\[\s*(\d+)s\]", l.strip())
            if m:
                dec_t.append(int(m.group(1)))
    dec = len(dec_t)
    done = bool(job.get("done"))
    t0 = job.get("t0")
    elapsed = (_time.monotonic() - t0) if t0 else 0
    total = pend
    # Médias medidas em runs reais: setup/descoberta ~55s, ~9s por GTO.
    SETUP_S, POR_GTO_S, BASELINE_GENERICO = 55, 9, 190
    est_total = (SETUP_S + POR_GTO_S * total) if total > 0 else BASELINE_GENERICO
    # Barra: acompanha o TEMPO (piso que sobe sempre) e pula à frente quando o
    # conteúdo real (downloads/decisões) avança mais rápido — nunca fica travada.
    if done:
        pct = 100
    else:
        time_pct = min(95, int(elapsed / max(est_total, 30) * 100))
        content_pct = int((baix + dec) / (2 * total) * 100) if total > 0 else 0
        pct = min(97, max(time_pct, content_pct))
    # ETA — sempre dá um número (sem "estimando"); fica mais preciso ao longo:
    #  1) com >=2 decisões: taxa REAL de decisões (a etapa lenta);
    #  2) descoberta feita (sabe o nº): baseline SETUP + nº*POR_GTO;
    #  3) setup/descoberta ainda rolando: baseline fixo genérico.
    eta = None
    if not done:
        if total > 0 and dec >= total:
            eta = 4  # tudo decidido, anexando/finalizando
        elif dec >= 2:
            rate = (dec_t[-1] - dec_t[0]) / (dec - 1)        # seg por decisão
            desde = max(0, elapsed - dec_t[-1])               # já passou desde a última
            eta = max(3, int((total - dec) * rate - desde))
        elif total > 0:                                       # baixando, sem decisão ainda
            eta = max(5, int(SETUP_S + POR_GTO_S * total - elapsed))
        else:                                                 # setup/descoberta
            eta = max(8, int(BASELINE_GENERICO - elapsed))
    if done:
        msg = "Concluído!"
    elif total == 0:
        msg = "Procurando GTOs pendentes do dia…"
    elif anex > 0:
        msg = f"Anexando no RedeUna… ({anex} enviado{'s' if anex != 1 else ''})"
    elif dec > 0:
        msg = f"Analisando solicitações com IA… ({dec} de {total})"
    elif baix > 0:
        msg = f"Baixando exames e laudos… ({baix} de {total})"
    else:
        msg = f"{total} pendentes encontrados — preparando…"
    return {"pct": pct, "msg": msg, "eta": eta, "total": total,
            "baixados": baix, "decididos": dec, "anexados": anex, "elapsed": int(elapsed)}


@app.route("/faturar/status/<jid>")
def faturar_status(jid: str):
    job = _esteira_jobs.get(jid)
    if not job:
        return jsonify({"error": "job não encontrado"}), 404
    r = job.get("resumo") or {}
    return jsonify({
        "done": job["done"], "error": job["error"],
        "execucao_id": job.get("execucao_id"),
        "prog": _esteira_progress(job),
        "resumo": {k: r.get(k) for k in ("solic_auto", "justificativa", "revisao",
                   "anexado_ok", "tempo_total", "pendentes")} if r else None,
    })


@app.route("/faturar/log/<jid>")
def faturar_log(jid: str):
    from flask import Response
    job = _esteira_jobs.get(jid)
    if not job:
        return "job não encontrado", 404
    head = f"# Log técnico — execução {jid} (dia {job.get('dia')})\n\n"
    return Response(head + "\n".join(job.get("log", [])),
                    mimetype="text/plain; charset=utf-8")


# ── APIs do Dashboard ─────────────────────────────────────────────────────────

@app.route("/api/dashboard")
def api_dashboard():
    """Dados agregados p/ o dashboard: última execução, semana, fila, totais."""
    try:
        with _jobs_lock:
            processando = sum(1 for j in _jobs.values()
                              if j.get("status") in ("running", "queued"))
            rodando_plano = {j.get("plano") for j in _jobs.values()
                             if j.get("status") in ("running", "queued")}
        ultima = db.run_mais_recente()
        # monta a lista de planos do registro + status (última execução de cada)
        por_plano = db.status_por_plano()
        lista_planos = []
        for p in planos_mod.listar_planos():
            lista_planos.append({
                "slug": p["slug"], "nome": p["nome"], "ativo": p.get("ativo", False),
                "rodando": p["slug"] in rodando_plano,
                "ultima": por_plano.get(p["slug"]),
            })
        return jsonify({
            "ultima": ultima,
            "planos": lista_planos,
            "recentes": db.ultimas_runs(8),
            "semana": db.serie_semana(),
            "revisao": db.fila_revisao(30),
            "totais": db.totais_gerais(),
            "processando": processando,
        })
    except Exception as exc:
        app.logger.error("Erro em /api/dashboard: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/api/gtos")
def api_gtos():
    """Funil + lista de GTOs de um dia (DD/MM/AAAA) ou da execução mais recente."""
    dia = request.args.get("dia", "").strip() or None
    try:
        run = db.run_mais_recente(dia)
        return jsonify({"run": run})
    except Exception as exc:
        app.logger.error("Erro em /api/gtos: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/api/planos-periodo")
def api_planos_periodo():
    """Gráfico empilhado: por plano, desfecho das GTOs no período [de, ate].
    `de`/`ate` em YYYY-MM-DD (o cliente calcula a partir de Hoje/Semana/Mês)."""
    de = request.args.get("de", "").strip()
    ate = request.args.get("ate", "").strip()
    try:
        agg = db.gtos_por_plano_periodo(de, ate) if de and ate else {}
        with _jobs_lock:
            rodando = {j.get("plano") for j in _jobs.values()
                       if j.get("status") in ("running", "queued")}
        planos = []
        for p in planos_mod.listar_planos():
            c = agg.get(p["slug"], {})
            planos.append({
                "slug": p["slug"], "nome": p["nome"], "ativo": p.get("ativo", False),
                "rodando": p["slug"] in rodando,
                "anexadas": c.get("anexadas", 0), "sem_laudo": c.get("sem_laudo", 0),
                "erros": c.get("erros", 0), "simulacao": c.get("simulacao", 0),
                "revisao": c.get("revisao", 0), "total": c.get("total", 0),
                "dias": c.get("dias", 0),
            })
        return jsonify({"de": de, "ate": ate, "planos": planos})
    except Exception as exc:
        app.logger.error("Erro em /api/planos-periodo: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/plano/<slug>")
def plano_detalhe_page(slug: str):
    """Tela de detalhe: GTOs processadas de um plano no período."""
    return render_template("plano_detalhe.html",
                           slug=slug, nome=planos_mod.nome_plano(slug),
                           de=request.args.get("de", ""), ate=request.args.get("ate", ""))


@app.route("/api/plano-detalhe")
def api_plano_detalhe():
    plano = request.args.get("plano", "").strip()
    de = request.args.get("de", "").strip()
    ate = request.args.get("ate", "").strip()
    if not plano or not de or not ate:
        return jsonify({"error": "Informe plano, de e ate."}), 400
    try:
        d = db.itens_plano_periodo(plano, de, ate)
        d["nome"] = planos_mod.nome_plano(plano)
        return jsonify(d)
    except Exception as exc:
        app.logger.error("Erro em /api/plano-detalhe: %s", exc)
        return jsonify({"error": str(exc)}), 500


# ── Panorama de Glosas / Recurso ──────────────────────────────────────────────

SITUACAO_META = {
    "A_RECORRER":           {"label": "A recorrer",  "cls": "warn",
                             "desc": "Glosada e ainda com recurso disponível na guia."},
    "RECURSO_OU_RESOLVIDA": {"label": "Recurso enviado / em análise", "cls": "info",
                             "desc": "Recursável, mas a guia já não mostra eventos glosados "
                                     "(recurso provavelmente já enviado ou resolvido)."},
    "RECURSO_REJEITADO":    {"label": "Recurso recusado (refazer)", "cls": "rej",
                             "desc": "Reanálise (recurso) feita de forma incorreta — já passou "
                                     "pelo recurso e foi recusada; precisa refazer."},
    "RESOLVIDA":            {"label": "Resolvida (paga)", "cls": "ok",
                             "desc": "Demonstrativo mostra a guia paga e sem glosa "
                                     "(recurso deferido ou glosa revertida)."},
    "GLOSA_CONFIRMADA":     {"label": "Glosa confirmada", "cls": "bad",
                             "desc": "Demonstrativo confirma a glosa no pagamento "
                                     "(recurso indeferido ou não recorrido)."},
    "NAO_RECURSAVEL":       {"label": "Não recursável", "cls": "bad",
                             "desc": "Recuperação de valores ou sem opção de recurso na guia."},
    "GLOSADA":              {"label": "Glosada", "cls": "neutral",
                             "desc": "Glosada (estado de recurso ainda não verificado)."},
}


def _glosa_view(lote: str = None) -> dict:
    pan = db.glosa_panorama(lote)
    evs = db.glosa_eventos(lote)
    pan["eventos"] = evs
    pan["meta"] = SITUACAO_META
    pan["lotes"] = db.glosa_lotes(12)
    return pan


@app.route("/glosas")
def glosas_page():
    """Panorama de glosas e recurso das unidades (tela, com exportar)."""
    return render_template("glosas.html")


@app.route("/api/glosas")
def api_glosas():
    try:
        return jsonify(_glosa_view(request.args.get("lote") or None))
    except Exception as exc:
        app.logger.error("Erro em /api/glosas: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/glosas/relatorio")
def glosas_relatorio():
    """Versão imprimível do panorama (usada também para gerar o PDF)."""
    pan = _glosa_view(request.args.get("lote") or None)
    embed = request.args.get("embed") in ("1", "true", "yes")
    return render_template("glosas_relatorio.html", pan=pan, pdf=False, embed=embed)


@app.route("/glosas.pdf")
def glosas_pdf():
    pan = _glosa_view(request.args.get("lote") or None)
    html = render_template("glosas_relatorio.html", pan=pan, pdf=True, embed=False)
    from playwright.sync_api import sync_playwright
    try:
        with sync_playwright() as pw:
            br = pw.chromium.launch(
                headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            pg = br.new_page()
            pg.set_content(html, wait_until="networkidle")
            pdf_bytes = pg.pdf(format="A4", print_background=True,
                               margin={"top": "12mm", "bottom": "12mm",
                                       "left": "10mm", "right": "10mm"})
            br.close()
    except Exception as exc:
        app.logger.error("Falha ao gerar PDF de glosas: %s", exc)
        return (f"Falha ao gerar PDF: {exc}", 500)
    nome = f"glosas_{(pan.get('dia') or '').replace('/', '-')}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name=nome)


@app.route("/glosas.xlsx")
def glosas_xlsx():
    import pandas as pd
    pan = _glosa_view(request.args.get("lote") or None)
    evs = pan.get("eventos", [])
    meta = SITUACAO_META
    # Resumo por unidade x situação
    resumo = []
    for u in pan.get("por_unidade", []):
        linha = {"Unidade": u["unidade"], "Total": u["total"]}
        for k, lbl in db.GLOSA_SITUACOES:
            linha[lbl] = u.get(k, 0)
        resumo.append(linha)
    det = [{
        "Unidade": e["unidade"], "Guia": e["ficha"], "Paciente": e["paciente"],
        "Procedimento": e["evento"], "Cód. glosa": e["glosa_cod"],
        "Motivo": e["glosa_motivo"],
        "Situação": meta.get(e["situacao"], {}).get("label", e["situacao"]),
    } for e in evs]
    motivos = [{"Cód.": m["glosa_cod"], "Motivo": m["glosa_motivo"], "Qtd": m["total"]}
               for m in pan.get("por_motivo", [])]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        (pd.DataFrame(resumo) if resumo else pd.DataFrame([{"Unidade": "—"}])).to_excel(
            xw, sheet_name="Resumo", index=False)
        (pd.DataFrame(motivos) if motivos else pd.DataFrame([{"Motivo": "—"}])).to_excel(
            xw, sheet_name="Por motivo", index=False)
        (pd.DataFrame(det) if det else pd.DataFrame([{"Guia": "—"}])).to_excel(
            xw, sheet_name="Glosas", index=False)
    buf.seek(0)
    nome = f"glosas_{(pan.get('dia') or '').replace('/', '-')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/glosas/atualizar", methods=["POST"])
def glosas_atualizar():
    """Dispara a atualização do panorama (background)."""
    body = request.get_json(silent=True) or {}
    dia = (body.get("dia") or datetime.now().strftime("%d/%m/%Y")).strip()
    checar = bool(body.get("checar_recurso", True))
    checar_demo = bool(body.get("checar_demonstrativo", True))
    contas = body.get("contas") or []
    job_id = uuid.uuid4().hex[:12]
    _purgar_jobs(_jobs, _jobs_lock)
    with _jobs_lock:
        _jobs[job_id] = {"status": "queued", "log": [], "kind": "glosa"}
    threading.Thread(target=_run_glosa_job, args=(job_id, dia, contas, checar, checar_demo),
                     daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/glosas/atualizar/status/<job_id>")
def glosas_atualizar_status(job_id: str):
    with _jobs_lock:
        j = _jobs.get(job_id)
        if not j:
            return jsonify({"error": "job não encontrado"}), 404
        return jsonify({"status": j.get("status"), "log": j.get("log", [])[-40:],
                        "total": j.get("total"), "lote": j.get("lote"),
                        "error": j.get("error")})


# ── Desfecho na RedeUna (pago/glosado/cancelado das guias que faturamos) ──────

DESFECHO_META = {
    "PAGA": {"label": "Paga", "cls": "ok", "desc": "Repasse processado e pago, sem glosa"},
    "GLOSADA": {"label": "Glosada", "cls": "bad", "desc": "A operadora recusou (ver motivo/recurso)"},
    "CANCELADA": {"label": "Cancelada", "cls": "neutral", "desc": "GTO cancelada ou não autorizada"},
    "AGUARDANDO": {"label": "Aguardando repasse", "cls": "info", "desc": "Ainda não processado no Demonstrativo"},
}


def _desfecho_view(lote: str = None) -> dict:
    pan = db.desfecho_panorama(lote)
    guias = db.desfechos(lote)
    return {"pan": pan, "guias": guias, "meta": DESFECHO_META}


@app.route("/desfecho")
def desfecho_page():
    """Status na RedeUna de cada guia que NÓS faturamos: pago/glosado/cancelado."""
    return render_template("desfecho.html")


@app.route("/api/desfecho")
def api_desfecho():
    try:
        return jsonify(_desfecho_view(request.args.get("lote") or None))
    except Exception as exc:
        app.logger.error("Erro em /api/desfecho: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/desfecho/atualizar", methods=["POST"])
def desfecho_atualizar():
    body = request.get_json(silent=True) or {}
    desde = (body.get("desde") or "").strip()
    if not desde:
        from datetime import date, timedelta
        desde = (date.today() - timedelta(days=120)).strftime("%d/%m/%Y")
    contas = body.get("contas") or []
    job_id = uuid.uuid4().hex[:12]
    _purgar_jobs(_jobs, _jobs_lock)
    with _jobs_lock:
        _jobs[job_id] = {"status": "queued", "log": [], "kind": "desfecho"}
    threading.Thread(target=_run_desfecho_job, args=(job_id, desde, contas),
                     daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/desfecho/atualizar/status/<job_id>")
def desfecho_atualizar_status(job_id: str):
    with _jobs_lock:
        j = _jobs.get(job_id)
        if not j:
            return jsonify({"error": "job não encontrado"}), 404
        return jsonify({"status": j.get("status"), "log": j.get("log", [])[-40:],
                        "total": j.get("total"), "lote": j.get("lote"),
                        "error": j.get("error")})


# ── Relatórios (hub) ──────────────────────────────────────────────────────────

def _relatorios_data() -> dict:
    """Junta faturamento (varredura) + glosas num pacote único p/ os relatórios."""
    anx = db.anexacao_panorama()
    glo = db.glosa_panorama()
    cat = anx.get("por_categoria", {}) or {}
    gsit = glo.get("glosado_situacao", {}) or {}
    sit = glo.get("por_situacao", {}) or {}
    # por unidade: cruza faturamento (anx) com glosado (glo)
    glo_uni = {u["unidade"]: u for u in glo.get("por_unidade", [])}
    por_unidade = []
    for u in anx.get("por_unidade", []):
        gu = glo_uni.get(u["unidade"], {})
        por_unidade.append({
            "unidade": u["unidade"], "total": u["total"],
            "faturada": u.get("FATURADA", 0), "a_faturar": u.get("A_FATURAR", 0),
            "sem_anexo": u.get("SEM_ANEXO", 0), "liberada": u.get("LIBERADA", 0),
            "cancelada": u.get("CANCELADA", 0),
            "glosas": gu.get("total", 0), "glosado": round(gu.get("glosado", 0), 2),
        })
    return {
        "anx": anx, "glo": glo,
        "resumo": {
            "total_gtos": anx.get("total", 0),
            "faturadas": cat.get("FATURADA", 0), "a_faturar": cat.get("A_FATURAR", 0),
            "sem_anexo": cat.get("SEM_ANEXO", 0), "liberadas": cat.get("LIBERADA", 0),
            "canceladas": cat.get("CANCELADA", 0),
            "glosas": glo.get("total", 0),
            "glosado_total": glo.get("total_glosado", 0),
            "recuperavel": round(gsit.get("A_RECORRER", 0), 2),
            "perda_confirmada": round(gsit.get("GLOSA_CONFIRMADA", 0), 2),
            "a_recorrer_n": sit.get("A_RECORRER", 0),
            "glosa_confirmada_n": sit.get("GLOSA_CONFIRMADA", 0),
        },
        "por_unidade": por_unidade,
        "pendencias": {
            "a_faturar": db.anexacao_gtos(categoria="A_FATURAR"),
            "a_recorrer": [e for e in db.glosa_eventos() if e["situacao"] == "A_RECORRER"],
        },
        "motivos_glosa": glo.get("por_motivo", []),
        "meta_sit": SITUACAO_META, "meta_cat": CATEGORIA_META,
    }


def _rel_dia_params():
    """Lê data (YYYY-MM-DD do <input type=date>) + contas do querystring.
    Devolve (dia_br, contas, data_iso, querystring)."""
    from config import PLANOS
    data_iso = (request.args.get("data") or "").strip()
    contas = [c for c in request.args.getlist("conta") if c in PLANOS]
    dia = ""
    if data_iso:
        try:
            y, m, d = data_iso.split("-")
            dia = f"{d}/{m}/{y}"
        except Exception:
            dia = ""
    qs = "&".join(["data=" + data_iso] + ["conta=" + c for c in contas])
    return dia, (contas or None), data_iso, qs


@app.route("/relatorios/dia")
def relatorios_dia():
    """Fechamento consolidado de um DIA (todas as unidades ou as escolhidas)."""
    from config import PLANOS
    dia, contas, data_iso, qs = _rel_dia_params()
    d = db.relatorio_dia(dia, contas) if dia else {"dia": "", "itens": [], "por_unidade": [],
                                                   "contas": contas or [],
                                                   "resumo": {"total": 0, "faturadas": 0,
                                                              "pendentes": 0, "por_categoria": {},
                                                              "execucoes": 0}}
    return render_template("relatorio_dia.html", d=d, planos=PLANOS,
                           data_iso=data_iso, qs=qs, pdf=False)


def _rel_pend_params():
    """Como _rel_dia_params, mas com data FINAL opcional (intervalo)."""
    from config import PLANOS
    dia, contas, data_iso, _qs = _rel_dia_params()
    fim_iso = (request.args.get("data_fim") or "").strip()
    dia_fim = ""
    if fim_iso:
        try:
            y, m, d = fim_iso.split("-")
            dia_fim = f"{d}/{m}/{y}"
        except Exception:
            dia_fim = ""
    qs = "&".join(["data=" + data_iso] + (["data_fim=" + fim_iso] if fim_iso else [])
                  + ["conta=" + c for c in (contas or [])])
    return dia, dia_fim, contas, data_iso, fim_iso, qs


@app.route("/relatorios/pendencias")
def relatorios_pendencias():
    """As pendências de um dia, agrupadas por QUEM precisa agir.

    O relatório do dia responde "o que não faturou e por quê". Este responde a
    pergunta seguinte, que é a que gera trabalho: "em qual delas eu mexo?".
    Numa lista de 10 nomes, 7 costumam se resolver sozinhas quando o laudo sair."""
    from config import PLANOS
    dia, dia_fim, contas, data_iso, fim_iso, qs = _rel_pend_params()
    d = db.pendencias_do_periodo(dia, dia_fim, contas) if dia else {
        "dia": "", "dias": [], "periodo": False, "grupos": [], "total": 0,
        "fila_tecnica": [], "total_fila": 0, "faturadas": 0, "total_guias": 0,
        "por_responsavel": {}, "por_unidade": [], "contas": contas or [],
        "unidades_fora": [], "dias_sem_execucao": []}
    # A tela promete que a fila tecnica "se resolve sozinha no reprocessamento".
    # Isso so e verdade com o cron LIGADO e dentro do prazo. Sem esses dois dados a
    # frase seria mentira exatamente quando mais importa — o cron vem desligado por
    # padrao. Melhor a tela contar a configuracao real do que uma promessa bonita.
    try:
        _prazo = int(os.environ.get("FATURAR_PRAZO_DIAS", "7"))
    except ValueError:
        _prazo = 7
    return render_template("relatorio_pendencias.html", d=d, planos=PLANOS,
                           data_iso=data_iso, fim_iso=fim_iso, qs=qs,
                           cron_ligado=(os.environ.get("FATURAR_CRON", "0") == "1"),
                           cron_hora=os.environ.get("FATURAR_CRON_HOUR", "5"),
                           prazo_dias=_prazo)


@app.route("/relatorios/pendencias.xlsx")
def relatorios_pendencias_xlsx():
    dia, dia_fim, contas, _iso, _fim, _qs = _rel_pend_params()
    if not dia:
        return ("Informe a data.", 400)
    d = db.pendencias_do_periodo(dia, dia_fim, contas)
    import pandas as pd
    linhas = []
    # SÓ o que é da operação. A fila técnica (falha nossa) fica de fora: este Excel
    # é entregue à clínica, e cobrar dela um bug nosso é pedir trabalho jogado fora.
    for g in d["grupos"]:
        for i in g["itens"]:
            linhas.append({"Quem resolve": g["responsavel"], "Situação": g["titulo"],
                           "Dia": i.get("dia") or d["dia"],
                           "Paciente": i["paciente"] or "—", "GTO": i["gto"],
                           "Unidade": i["unidade"],
                           "Exames (GTO)": i["exames_gto"] or "—",
                           "O que fazer": g["acao"],
                           "Motivo completo": i["motivo"] or "—"})
    df = pd.DataFrame(linhas)
    df_r = pd.DataFrame([{"Quem resolve": k, "Pendências": v}
                         for k, v in sorted(d["por_responsavel"].items(),
                                            key=lambda x: -x[1])])
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        (df_r if not df_r.empty else pd.DataFrame([{"Quem resolve": "(nenhuma pendência)"}])
         ).to_excel(xw, sheet_name="Resumo", index=False)
        (df if not df.empty else pd.DataFrame([{"GTO": "(nenhuma)"}])
         ).to_excel(xw, sheet_name="Pendencias", index=False)
        from openpyxl.styles import Font, PatternFill, Alignment
        for nome_ws in ("Resumo", "Pendencias"):
            ws = xw.book[nome_ws]
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="0F7A4F")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[1].height = 24
            ws.freeze_panes = "A2"
            for col in ws.columns:
                largura = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(largura + 3, 12), 70)
    bio.seek(0)
    _suf = dia.replace("/", "-") + (f"_a_{dia_fim.replace('/', '-')}"
                                    if dia_fim and dia_fim != dia else "")
    nome = f"pendencias_{_suf}.xlsx"
    return send_file(bio, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument."
                              "spreadsheetml.sheet")


@app.route("/relatorios/dia.pdf")
def relatorios_dia_pdf():
    from config import PLANOS
    dia, contas, data_iso, qs = _rel_dia_params()
    if not dia:
        return ("Informe a data.", 400)
    d = db.relatorio_dia(dia, contas)
    html = render_template("relatorio_dia.html", d=d, planos=PLANOS,
                           data_iso=data_iso, qs=qs, pdf=True)
    from playwright.sync_api import sync_playwright
    try:
        with sync_playwright() as pw:
            br = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            pg = br.new_page()
            pg.set_content(html, wait_until="networkidle")
            pdf_bytes = pg.pdf(format="A4", print_background=True,
                               margin={"top": "12mm", "bottom": "12mm",
                                       "left": "10mm", "right": "10mm"})
            br.close()
    except Exception as exc:
        return (f"Falha ao gerar PDF: {exc}", 500)
    nome = f"relatorio_{dia.replace('/', '-')}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name=nome)


@app.route("/relatorios/dia.xlsx")
def relatorios_dia_xlsx():
    dia, contas, _iso, _qs = _rel_dia_params()
    if not dia:
        return ("Informe a data.", 400)
    d = db.relatorio_dia(dia, contas)
    import pandas as pd
    fat = [i for i in d["itens"] if i["faturado"]]
    pen = [i for i in d["itens"] if not i["faturado"]]
    df_f = pd.DataFrame([{"Paciente": i["paciente"] or "—", "GTO": i["gto"],
                          "Unidade": i["unidade"],
                          "Exames (GTO)": i["exames_gto"] or "—",
                          "Documento anexado": (
                              i["solicitacao"]
                              or ("justificativa na GTO (campo 49) — solicitação dispensada"
                                  if i.get("categoria") == "justificativa"
                                  else "anexado em execução anterior"))}
                         for i in fat])
    df_p = pd.DataFrame([{"Paciente": i["paciente"] or "—", "GTO": i["gto"],
                          "Unidade": i["unidade"],
                          "Exames (GTO)": i["exames_gto"] or "—",
                          "Motivo": i["motivo"] or (i["categoria"] or "").replace("_", " ") or "—"}
                         for i in pen])
    df_u = pd.DataFrame([{"Unidade": u["unidade"], "Total": u["total"],
                          "Faturadas": u["faturadas"], "Pendentes": u["pendentes"]}
                         for u in d["por_unidade"]])
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        (df_u if not df_u.empty else pd.DataFrame([{"Unidade": "(sem dados)"}])
         ).to_excel(xw, sheet_name="Resumo", index=False)
        (df_f if not df_f.empty else pd.DataFrame([{"GTO": "(nenhuma)"}])
         ).to_excel(xw, sheet_name="Faturadas", index=False)
        (df_p if not df_p.empty else pd.DataFrame([{"GTO": "(nenhuma)"}])
         ).to_excel(xw, sheet_name="Pendentes", index=False)
        for nome_ws in ("Resumo", "Faturadas", "Pendentes"):
            ws = xw.book[nome_ws]
            from openpyxl.styles import Font, PatternFill, Alignment
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="0F7A4F")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[1].height = 24
            ws.freeze_panes = "A2"
            for col in ws.columns:
                largura = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(largura + 3, 12), 60)
    bio.seek(0)
    nome = f"relatorio_{dia.replace('/', '-')}.xlsx"
    return send_file(bio, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/relatorios")
def relatorios_page():
    return render_template("relatorios.html")


@app.route("/api/relatorios")
def api_relatorios():
    try:
        return jsonify(_relatorios_data())
    except Exception as exc:
        app.logger.error("Erro em /api/relatorios: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/relatorios.pdf")
def relatorios_pdf():
    data = _relatorios_data()
    html = render_template("relatorios_pdf.html", d=data)
    from playwright.sync_api import sync_playwright
    try:
        with sync_playwright() as pw:
            br = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
            pg = br.new_page()
            pg.set_content(html, wait_until="networkidle")
            pdf_bytes = pg.pdf(format="A4", print_background=True,
                               margin={"top": "12mm", "bottom": "12mm", "left": "10mm", "right": "10mm"})
            br.close()
    except Exception as exc:
        return (f"Falha ao gerar PDF: {exc}", 500)
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                     as_attachment=True, download_name="relatorios_radiobras.pdf")


# ── Anexação / Faturamento (varredura só-leitura) ─────────────────────────────

CATEGORIA_META = {
    "FATURADA":  {"label": "Faturada", "cls": "ok",
                  "desc": "2+ anexos — laudo e entrega já anexados"},
    "A_FATURAR": {"label": "A faturar", "cls": "warn",
                  "desc": "Só 1 anexo — falta o 2º para faturar"},
    "SEM_ANEXO": {"label": "Sem anexo", "cls": "bad",
                  "desc": "Nenhum anexo enviado ainda"},
    "LIBERADA":  {"label": "Liberada p/ assinatura", "cls": "info",
                  "desc": "Senha liberada no portal — aguardando"},
    "CANCELADA": {"label": "Cancelada", "cls": "neutral",
                  "desc": "GTO cancelada ou não autorizada"},
    "ERRO":      {"label": "Erro de leitura", "cls": "bad",
                  "desc": "Não foi possível ler a GTO"},
}


def _anexacao_view(lote: str = None) -> dict:
    pan = db.anexacao_panorama(lote)
    pan["gtos"] = db.anexacao_gtos(lote)
    pan["meta"] = CATEGORIA_META
    pan["lotes"] = db.anexacao_lotes(12)
    return pan


@app.route("/api/anexacao")
def api_anexacao():
    try:
        return jsonify(_anexacao_view(request.args.get("lote") or None))
    except Exception as exc:
        app.logger.error("Erro em /api/anexacao: %s", exc)
        return jsonify({"error": str(exc)}), 500


@app.route("/anexacao/atualizar", methods=["POST"])
def anexacao_atualizar():
    body = request.get_json(silent=True) or {}
    hoje = datetime.now().strftime("%d/%m/%Y")
    de = (body.get("de") or ("01/" + hoje[3:])).strip()
    ate = (body.get("ate") or hoje).strip()
    contas = body.get("contas") or []
    limite = int(body.get("limite") or 0)
    job_id = uuid.uuid4().hex[:12]
    _purgar_jobs(_jobs, _jobs_lock)
    with _jobs_lock:
        _jobs[job_id] = {"status": "queued", "log": [], "kind": "anexacao"}
    threading.Thread(target=_run_anexacao_job, args=(job_id, de, ate, contas, limite),
                     daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/anexacao/atualizar/status/<job_id>")
def anexacao_atualizar_status(job_id: str):
    with _jobs_lock:
        j = _jobs.get(job_id)
        if not j:
            return jsonify({"error": "job não encontrado"}), 404
        return jsonify({"status": j.get("status"), "log": j.get("log", [])[-40:],
                        "total": j.get("total"), "lote": j.get("lote"),
                        "error": j.get("error")})


@app.route("/api/diag")
def api_diag():
    """Diagnóstico de saúde — para inspeção remota sem acesso aos logs do servidor.
    Reporta: banco conectado, presença de credenciais e últimas execuções (c/ erro)."""
    from sqlalchemy import text
    diag = {"app": "ok"}
    # banco
    try:
        with db.engine.connect() as c:
            c.execute(text("SELECT 1"))
        diag["db_ok"] = True
        diag["db_host"] = db.DATABASE_URL.split("@")[-1]
    except Exception as e:
        diag["db_ok"] = False
        diag["db_error"] = str(e)[:200]
    # credenciais presentes? (não expõe valores)
    diag["cred"] = {
        "smartris": bool(os.environ.get("SMARTRIS_EMAIL") and os.environ.get("SMARTRIS_PASSWORD")),
        "odontoprev": bool(os.environ.get("ODONTOPREV_USER") and os.environ.get("ODONTOPREV_PASSWORD")),
    }
    # jobs em memória + últimas execuções (com erro resumido)
    # ATENÇÃO: contava só `_jobs` (fluxo antigo). As execuções de /faturar vivem em
    # `_esteira_jobs` e ficavam de fora — o diagnóstico dizia "0 jobs ativos" com a
    # esteira faturando. Como é por ele que se decide se pode deployar, e deploy no
    # meio de uma execução MATA o job, a resposta errada aqui custa um faturamento.
    with _jobs_lock:
        _antigos = sum(1 for j in _jobs.values()
                       if j.get("status") in ("running", "queued"))
    _esteira_ativa = [{"job": k, "dia": (j.get("data") or j.get("dia")),
                       "conta": j.get("plano") or j.get("conta"),
                       "dry": j.get("dry")}
                      for k, j in list(_esteira_jobs.items()) if not j.get("done")]
    diag["jobs_ativos"] = _antigos + len(_esteira_ativa)
    diag["esteira_ativa"] = _esteira_ativa
    diag["pode_deployar"] = (diag["jobs_ativos"] == 0)
    with _esteira_ativas_lock:
        diag["travas"] = [{"dia": d, "conta": c,
                           "ha_segundos": round(time.monotonic() - v["t"])}
                          for (d, c), v in _esteira_ativas.items()]
    try:
        diag["runs"] = [{
            "id": r["id"], "plano": r["plano"], "dia": r["dia"], "status": r["status"],
            "enviados": r["enviados"], "erros": r["erros"],
            "erro_msg": (r.get("erro_msg") or "")[:300],
            "finished_at": r["finished_at"],
        } for r in db.runs_recentes(10)]
    except Exception as e:
        diag["runs_error"] = str(e)[:200]
    # faturamento (fluxo atual): estado do cron D-4, backlog e últimas execuções
    try:
        ult = db.cron_faturar_last_at()
        diag["faturamento"] = {
            "cron_ligado": os.environ.get("FATURAR_CRON", "0") == "1",
            "cron_ultima": ult.isoformat() if ult else None,
            "cron_rodou_hoje": _faturar_rodou_hoje(),
            "prazo_dias": _prazo_dias(),
            "pendencias_abertas": db.contar_pendencias_abertas(),
            "alerta_sla_ligado": os.environ.get("ALERTA_SLA", "1") != "0",
            "smtp_configurado": bool(os.environ.get("SMTP_HOST") and os.environ.get("ALERTA_EMAIL_TO")),
            # canal de falha tecnica (22/08). Sem isso a tela prometeria um aviso que
            # nunca sai — o mesmo erro do SMTP dormente descoberto em 02/08.
            "whatsapp_configurado": _whatsapp_ok(),
            "alerta_falha_ligado": os.environ.get("ALERTA_FALHA", "1") != "0",
            "retry_cron": os.environ.get("RETRY_CRON", "0") == "1",
            # disjuntor: se a fila esta parada por falha global, a tela tem que dizer
            # — senao "0 retries" parece calmaria quando e apagao.
            "retry_pausado": db.retry_pausado(),
            "retry_pausa": (lambda i: {"ate": i["ate"].isoformat() if i.get("ate") else None,
                                       "motivo": (i.get("motivo") or "")[:200],
                                       "ativa": i.get("ativa")})(db.retry_pausa_info() or {})
            if db.retry_pausa_info() else None,
            "faturar_cron": os.environ.get("FATURAR_CRON", "0") == "1",
            "faturar_cron_hora": os.environ.get("FATURAR_CRON_HOUR", "5"),
            "faturar_prazo_dias": os.environ.get("FATURAR_PRAZO_DIAS", "7"),
            "execucoes": [{
                "id": e["id"], "dia": e["dia"], "conta": e["conta"],
                "dry_run": e["dry_run"], "faturadas": e["faturadas"],
                "pendentes": e["pendentes"], "criado_em": (
                    e["criado_em"].isoformat() if e.get("criado_em") else None),
            } for e in db.listar_execucoes(5)],
        }
    except Exception as e:
        diag["faturamento_error"] = str(e)[:200]
    return jsonify(diag)


@app.route("/fechar", methods=["POST"])
def fechar_route():
    """Inicia o fechamento do dia (download + anexo no OdontoPrev). Assíncrono."""
    data = request.form.get("data", "").strip()
    plano = (request.form.get("plano", "") or "odontoprev").strip()
    # 'simular' = dry-run (não anexa, só mostra o que faria)
    dry_run = request.form.get("simular", "").lower() in ("1", "true", "on", "yes")
    if not data:
        return jsonify({"error": "Informe o dia (DD/MM/AAAA)."}), 400
    if not re.match(r"^\d{2}/\d{2}/\d{4}$", data):
        return jsonify({"error": "Data inválida. Use DD/MM/AAAA."}), 400
    if not planos_mod.plano_ativo(plano):
        return jsonify({"error": f"O plano '{planos_mod.nome_plano(plano)}' ainda não "
                                 "está configurado para automação."}), 400

    job_id = str(uuid.uuid4())[:8]
    _purgar_jobs(_jobs, _jobs_lock)
    with _jobs_lock:
        _jobs[job_id] = {"status": "queued", "log": [], "plano": plano}
    threading.Thread(
        target=_run_fechar_job, args=(job_id, data, dry_run, plano), daemon=True
    ).start()
    return jsonify({"job_id": job_id, "dry_run": dry_run})


@app.route("/fechar/status/<job_id>")
def fechar_status(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job não encontrado."}), 404
    resp: dict = {"status": job["status"], "log": job.get("log", [])}
    resp["run_id"] = job.get("run_id")
    if job["status"] == "error":
        # NÃO vaza o traceback/erro técnico pro usuário — só uma mensagem amigável
        # + um código (run_id) p/ a equipe achar os detalhes nos logs (DB/diag).
        resp["error"] = "Não foi possível concluir o processamento."
        resp["error_code"] = str(job.get("run_id") or job_id)
    if job["status"] == "done":
        rel = job.get("relatorio", {})
        resp["resumo"] = rel.get("resumo", {})
        resp["itens"] = rel.get("itens", [])
        resp["dry_run"] = rel.get("dry_run", False)
    return jsonify(resp)


@app.route("/gerar", methods=["POST"])
def gerar():
    date_from = request.form.get("date_from", "").strip()
    date_to = request.form.get("date_to", "").strip()
    selected_convenios = request.form.getlist("convenios")
    selected_segmentos = request.form.getlist("segmentos")

    if not date_from or not date_to:
        return jsonify({"error": "Informe o período."}), 400
    if not selected_convenios:
        return jsonify({"error": "Selecione ao menos um convênio."}), 400
    if not selected_segmentos:
        return jsonify({"error": "Selecione ao menos um segmento."}), 400

    try:
        email, password = get_credentials()
        convenio_map, segmento_map, cookies = discover_tokens_and_cookies(email, password)
        insurance_tokens = resolve_tokens(selected_convenios, convenio_map, "convenio")
        segment_tokens = resolve_tokens(selected_segmentos, segmento_map, "segmento")

        if not insurance_tokens:
            return jsonify({"error": "Nenhum convênio resolvido. Verifique os nomes."}), 400

        html = post_relatorio(cookies, insurance_tokens, segment_tokens, date_from, date_to)
        df, valor_total, num_exames = parse_html_to_df(html)

        if df.empty:
            return jsonify({"warning": "Nenhum exame encontrado para o período."}), 200

        buf = io.BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        date_tag = date_from.replace("/", "") + "_" + date_to.replace("/", "")
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"pacientes_analitico_REDEUNNA_{date_tag}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as exc:
        tb = traceback.format_exc()
        app.logger.error("Erro em /gerar:\n%s", tb)
        return jsonify({"error": str(exc), "traceback": tb}), 500


@app.route("/baixar_dia", methods=["POST"])
def baixar_dia():
    date_from = request.form.get("date_from", "").strip()
    date_to = request.form.get("date_to", "").strip()
    selected_convenios = request.form.getlist("convenios")
    selected_segmentos = request.form.getlist("segmentos")

    if not date_from or not date_to:
        return jsonify({"error": "Informe o período."}), 400
    if not selected_convenios:
        return jsonify({"error": "Selecione ao menos um convênio."}), 400

    # Usar date_from como data do dia (dia único)
    job_id = str(uuid.uuid4())[:8]
    _purgar_jobs(_jobs, _jobs_lock)
    with _jobs_lock:
        _jobs[job_id] = {"status": "queued", "log": []}

    thread = threading.Thread(
        target=_run_job,
        args=(job_id, date_from, selected_convenios, selected_segmentos),
        daemon=True,
    )
    thread.start()
    return jsonify({"job_id": job_id})


@app.route("/baixar_dia/status/<job_id>")
def baixar_dia_status(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job não encontrado."}), 404

    resp: dict = {"status": job["status"], "log": job.get("log", [])}
    if job["status"] == "error":
        resp["error"] = job.get("error", "")
    if job["status"] == "done":
        resp["resumo"] = job.get("relatorio", {}).get("resumo", {})
    return jsonify(resp)


@app.route("/baixar_dia/resultado/<job_id>")
def baixar_dia_resultado(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job não encontrado."}), 404
    if job["status"] != "done":
        return jsonify({"error": "Job ainda não concluído."}), 400

    zpath = job.get("zip_path")
    if not zpath or not os.path.exists(zpath):
        return jsonify({"error": "Arquivo expirado — rode a extração de novo."}), 410
    data_tag = (
        job.get("relatorio", {}).get("periodo", {}).get("de", "").replace("/", "")
    )
    filename = f"arquivos_REDEUNNA_{data_tag}.zip"
    return send_file(zpath, as_attachment=True, download_name=filename,
                     mimetype="application/zip")


@app.route("/ciclo_dia", methods=["POST"])
def ciclo_dia_route():
    # DESATIVADA em 25/07 pela auditoria. Esta rota ANEXAVA DE VERDADE (a chamada
    # não passava dry_run e o padrão de ciclo_dia() é False), com um clique, para
    # qualquer usuário logado — e SEM NENHUMA das guardas que o pipeline principal
    # tem: não lia os exames da GTO, não filtrava exame PARTICULAR (subia a pasta
    # inteira), não conferia o nº da guia nem o nome do paciente, não exigia
    # solicitação/justificativa e não gravava nada no banco.
    # ciclo_completo.py não é alterado desde 12/06 — nenhuma correção chegou nele.
    # É redundante: /faturar (esteira) e FECHAR DIA fazem o mesmo COM as guardas.
    return jsonify({
        "error": "O 'Ciclo Completo' foi desativado por segurança: ele anexava sem "
                 "as verificações de paciente, laudo e exame particular. Use "
                 "'Faturar dia' (/faturar), que faz o mesmo com todas as guardas."
    }), 410

    date_from = request.form.get("date_from", "").strip()
    selected_convenios = request.form.getlist("convenios") or CONVENIOS
    selected_segmentos = request.form.getlist("segmentos") or SEGMENTOS

    if not date_from:
        return jsonify({"error": "Informe o dia."}), 400

    job_id = str(uuid.uuid4())[:8]
    _purgar_jobs(_jobs, _jobs_lock)
    with _jobs_lock:
        _jobs[job_id] = {"status": "queued", "log": []}

    threading.Thread(
        target=_run_ciclo_job,
        args=(job_id, date_from, selected_convenios, selected_segmentos),
        daemon=True,
    ).start()
    return jsonify({"job_id": job_id})


@app.route("/ciclo_dia/status/<job_id>")
def ciclo_dia_status(job_id: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job não encontrado."}), 404
    resp: dict = {"status": job["status"], "log": job.get("log", [])}
    if job["status"] == "error":
        resp["error"] = job.get("error", "")
    if job["status"] == "done":
        resp["relatorio"] = job.get("relatorio", {})
    return jsonify(resp)


# ── Job store da esteira (/faturar) ───────────────────────────────────────────
# As rotas /admin/esteira/* foram REMOVIDAS em 27/07 (auditoria). Eram a mesma
# porta dos fundos fechada no /ciclo_dia: GET (acionável por link/prefetch), chave
# com default `rb-esteira-2026` COMMITADO em repositório público, sem checagem de
# admin, aceitando dry=0 (anexação REAL) e sem receber `conta` — caía no
# ODONTOPREV_USER padrão, ou seja, faturava na UNIDADE ERRADA, justamente o que o
# rodar_esteira passou a rejeitar com ValueError.
# A tela de revisão que vivia ali servia documento de paciente autenticado por
# chave na URL (vaza em log de proxy, histórico e print). Tudo que faziam tem
# substituto: /faturar dispara, /pendencias mostra o backlog, /faturar/log/<jid> traz
# o log técnico.
_esteira_jobs: dict = {}


if __name__ == "__main__":
    # Dev local. Em produção (Docker/EasyPanel) o servidor é o gunicorn.
    port = int(os.environ.get("PORT", "5000"))
    app.run(debug=False, host="0.0.0.0", port=port)
